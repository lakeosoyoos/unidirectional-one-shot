"""
reburn_percentage.py — Reburn metric for the Excel workbook
===========================================================

Calculates how much of the cable needs a re-splice, expressed as a
fraction of the ribbon × splice grid:

    reburn % = (# of ribbon × splice cells with ≥ 1 fiber flagged for
                reburn at that splice)
              ─────────────────────────────────────────────────────
              (# of ribbons × # of splice columns)

Worked example (from the spec the maintainer gave):
    70 ribbons × 10 splices  =  700 total cells
    14 ribbons each had 5 splices with at least one fiber needing a
    reburn                   =   70 reburn cells
    70 / 700 = 10 %

What counts:
    • Only SPLICE columns (kind == 'splice').  Bend/Damage and Break
      columns describe other problems and are excluded.
    • A cell counts as "needing a reburn" if ANY fiber in that ribbon
      hits the splice with |loss| >= BEND_THRESHOLD (= 0.100 dB).  Since
      build_ribbon_grid already applies that threshold, ANY non-empty
      cell in a splice column is a reburn cell.

Sheet position: inserted at index 1, immediately after the
"Acquisition Parameters" sheet, so the workbook still opens on the
audit sheet but the reburn metric is the next thing the tech sees.
"""
from __future__ import annotations

from typing import Iterable

try:
    import openpyxl                         # type: ignore
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:                         # pragma: no cover
    openpyxl = None


# ─────────────────────────────────────────────────────────────────────
#  Summary builder
# ─────────────────────────────────────────────────────────────────────

def build_reburn_summary(grid: dict, columns: list, n_ribbons: int,
                         ribbon_label_fn=None) -> dict:
    """Compute the reburn metric + breakdowns.

    Returns::

        {
          'n_ribbons':       int,
          'n_splice_cols':   int,
          'total_cells':     int,
          'reburn_cells':    int,
          'percentage':      float,           # 0..100
          'per_splice':      [ { 'splice_label': str,
                                  'km':           float,
                                  'n_ribbons':    int,
                                  'pct':          float }, ... ],
          'per_ribbon':      [ { 'ribbon_label': str,
                                  'ribbon_idx':   int,
                                  'n_splices':    int,
                                  'pct':          float }, ... ],
        }

    ``grid`` is the {(ribbon_idx, col_idx): [(fiber, loss), ...]} map
    produced by unidirectional_event_finder.build_ribbon_grid.
    ``columns`` is the column-metadata list in the same order
    write_xlsx renders them.  ``ribbon_label_fn`` is an optional
    ``(ribbon_idx) -> str`` callable used to label the per-ribbon
    breakdown rows; falls back to ``"Ribbon N"``.
    """
    if ribbon_label_fn is None:
        ribbon_label_fn = lambda ri: f"Ribbon {ri + 1}"

    # Only splice columns participate in the reburn calculation.
    splice_cols = [(ci, col) for ci, col in enumerate(columns)
                   if col.get('kind') == 'splice']
    n_splice_cols = len(splice_cols)
    total_cells = n_ribbons * n_splice_cols

    # ── Aggregate counts ───────────────────────────────────────────
    reburn_cells = 0
    per_splice_counts = [0] * n_splice_cols
    per_ribbon_counts = [0] * n_ribbons

    for split_idx, (ci, col) in enumerate(splice_cols):
        for ri in range(n_ribbons):
            if grid.get((ri, ci)):
                reburn_cells     += 1
                per_splice_counts[split_idx] += 1
                per_ribbon_counts[ri]        += 1

    percentage = (reburn_cells / total_cells * 100.0) if total_cells else 0.0

    # ── Per-splice rows, in distance order (same as the grid sheet) ─
    per_splice = []
    for split_idx, (ci, col) in enumerate(splice_cols):
        per_splice.append({
            'splice_label': f"Splice {split_idx + 1}",
            'km':           float(col.get('position_km_display',
                                          col.get('position_km_refined', 0.0))),
            'n_ribbons':    per_splice_counts[split_idx],
            'pct':          (per_splice_counts[split_idx] / n_ribbons * 100.0)
                            if n_ribbons else 0.0,
        })

    # ── Per-ribbon rows ─────────────────────────────────────────────
    per_ribbon = []
    for ri in range(n_ribbons):
        per_ribbon.append({
            'ribbon_label': ribbon_label_fn(ri),
            'ribbon_idx':   ri,
            'n_splices':    per_ribbon_counts[ri],
            'pct':          (per_ribbon_counts[ri] / n_splice_cols * 100.0)
                            if n_splice_cols else 0.0,
        })

    return {
        'n_ribbons':     n_ribbons,
        'n_splice_cols': n_splice_cols,
        'total_cells':   total_cells,
        'reburn_cells':  reburn_cells,
        'percentage':    percentage,
        'per_splice':    per_splice,
        'per_ribbon':    per_ribbon,
    }


# ─────────────────────────────────────────────────────────────────────
#  Excel writer
# ─────────────────────────────────────────────────────────────────────

_FONT_NAME = "Calibri"
_FONT_SIZE = 12


def _styles():
    return {
        'hdr_font':    Font(name=_FONT_NAME, size=_FONT_SIZE,
                            bold=True, color="FFFFFF"),
        'hdr_fill':    PatternFill(start_color="1F4E79",
                                    end_color="1F4E79",
                                    fill_type="solid"),
        'body':        Font(name=_FONT_NAME, size=_FONT_SIZE),
        'body_bold':   Font(name=_FONT_NAME, size=_FONT_SIZE, bold=True),
        'big_pct':     Font(name=_FONT_NAME, size=28, bold=True,
                            color="9C5700"),
        'big_pct_lbl': Font(name=_FONT_NAME, size=12, color="9C5700",
                            italic=True),
        'subhdr_fill': PatternFill(start_color="DDEBF7",
                                    end_color="DDEBF7",
                                    fill_type="solid"),
        'good_fill':   PatternFill(start_color="C6EFCE",
                                    end_color="C6EFCE",
                                    fill_type="solid"),
        'warn_fill':   PatternFill(start_color="FFEB9C",
                                    end_color="FFEB9C",
                                    fill_type="solid"),
        'border':      Border(left=Side(style='thin', color='CCCCCC'),
                              right=Side(style='thin', color='CCCCCC'),
                              top=Side(style='thin', color='CCCCCC'),
                              bottom=Side(style='thin', color='CCCCCC')),
    }


def write_reburn_sheet(wb, summary: dict, insert_at: int = 1,
                       sheet_title: str = "Reburn Percentage") -> None:
    """Insert the Reburn Percentage sheet at ``insert_at`` in ``wb``.

    Idempotent: if a sheet by that name already exists it's removed
    first so re-runs produce a fresh layout.  Default ``insert_at=1``
    places the sheet immediately after the Acquisition Parameters
    sheet (sheet 0) so the audit remains the workbook's active sheet
    on open.
    """
    if openpyxl is None:
        return
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]

    ws = wb.create_sheet(title=sheet_title, index=insert_at)
    ws.sheet_properties.tabColor = "C0185F"

    S = _styles()

    # ── Title ───────────────────────────────────────────────────────
    title = ws.cell(row=1, column=1, value="Reburn percentage")
    title.font = Font(name=_FONT_NAME, size=14, bold=True, color="FFFFFF")
    title.fill = S['hdr_fill']
    ws.cell(row=1, column=2).fill = S['hdr_fill']
    ws.cell(row=1, column=3).fill = S['hdr_fill']
    ws.cell(row=1, column=4).fill = S['hdr_fill']

    sub = ws.cell(row=2, column=1,
                  value="Splice cells that need a reburn, "
                        "expressed as a fraction of the ribbon × splice grid.")
    sub.font = Font(name=_FONT_NAME, size=10, italic=True, color="595959")

    # ── Big number ──────────────────────────────────────────────────
    pct = summary['percentage']
    pct_cell = ws.cell(row=4, column=1, value=f"{pct:.2f}%")
    pct_cell.font      = S['big_pct']
    pct_cell.alignment = Alignment(horizontal='left', vertical='center')
    pct_lbl = ws.cell(row=4, column=2,
                      value=f"{summary['reburn_cells']:,} of "
                            f"{summary['total_cells']:,} splice cells")
    pct_lbl.font = S['big_pct_lbl']
    pct_lbl.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[4].height = 38

    # ── Calculation breakdown table ────────────────────────────────
    row = 6
    h = ws.cell(row=row, column=1, value="The calculation")
    h.font = S['hdr_font']; h.fill = S['hdr_fill']
    ws.cell(row=row, column=2).fill = S['hdr_fill']
    row += 1

    calc_rows = [
        ("Ribbons",                       summary['n_ribbons']),
        ("Splice columns",                summary['n_splice_cols']),
        ("Total ribbon × splice cells",   summary['total_cells']),
        ("Cells with ≥ 1 fiber needing reburn",
                                          summary['reburn_cells']),
        ("Reburn percentage",
                                          f"{summary['percentage']:.2f}%"),
    ]
    for label, value in calc_rows:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = S['body_bold']; lc.border = S['border']
        lc.alignment = Alignment(vertical='center')
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = S['body']; vc.border = S['border']
        vc.alignment = Alignment(vertical='center')
        row += 1

    # ── Per-splice breakdown ───────────────────────────────────────
    row += 2
    h = ws.cell(row=row, column=1, value="Per-splice breakdown")
    h.font = S['hdr_font']; h.fill = S['hdr_fill']
    for c in (2, 3, 4):
        ws.cell(row=row, column=c).fill = S['hdr_fill']
    row += 1

    headers = ("Splice", "Distance (km)", "Ribbons needing reburn",
               "Percentage of ribbons")
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=True,
                      color="1F4E79")
        c.fill = S['subhdr_fill']; c.border = S['border']
        c.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    for sp in summary['per_splice']:
        lcell = ws.cell(row=row, column=1, value=sp['splice_label'])
        lcell.font = S['body']; lcell.border = S['border']
        kcell = ws.cell(row=row, column=2, value=round(sp['km'], 2))
        kcell.font = S['body']; kcell.border = S['border']
        ncell = ws.cell(row=row, column=3,
                        value=f"{sp['n_ribbons']} of {summary['n_ribbons']}")
        ncell.font = S['body']; ncell.border = S['border']
        ncell.alignment = Alignment(horizontal='center')
        pcell = ws.cell(row=row, column=4, value=f"{sp['pct']:.1f}%")
        pcell.font = S['body']; pcell.border = S['border']
        pcell.alignment = Alignment(horizontal='center')
        # Light shading: green if 0%, amber if any.
        if sp['n_ribbons'] == 0:
            for c in (1, 2, 3, 4):
                ws.cell(row=row, column=c).fill = S['good_fill']
        else:
            for c in (1, 2, 3, 4):
                ws.cell(row=row, column=c).fill = S['warn_fill']
        row += 1

    # ── Per-ribbon breakdown ───────────────────────────────────────
    row += 2
    h = ws.cell(row=row, column=1, value="Per-ribbon breakdown")
    h.font = S['hdr_font']; h.fill = S['hdr_fill']
    for c in (2, 3, 4):
        ws.cell(row=row, column=c).fill = S['hdr_fill']
    row += 1

    headers = ("Ribbon", "Splices needing reburn", "of total splices",
               "Percentage")
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=True,
                      color="1F4E79")
        c.fill = S['subhdr_fill']; c.border = S['border']
        c.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    for rb in summary['per_ribbon']:
        lcell = ws.cell(row=row, column=1, value=rb['ribbon_label'])
        lcell.font = S['body']; lcell.border = S['border']
        ncell = ws.cell(row=row, column=2, value=rb['n_splices'])
        ncell.font = S['body']; ncell.border = S['border']
        ncell.alignment = Alignment(horizontal='center')
        tcell = ws.cell(row=row, column=3, value=summary['n_splice_cols'])
        tcell.font = S['body']; tcell.border = S['border']
        tcell.alignment = Alignment(horizontal='center')
        pcell = ws.cell(row=row, column=4, value=f"{rb['pct']:.1f}%")
        pcell.font = S['body']; pcell.border = S['border']
        pcell.alignment = Alignment(horizontal='center')
        if rb['n_splices'] == 0:
            for c in (1, 2, 3, 4):
                ws.cell(row=row, column=c).fill = S['good_fill']
        else:
            for c in (1, 2, 3, 4):
                ws.cell(row=row, column=c).fill = S['warn_fill']
        row += 1

    # ── Column widths ──────────────────────────────────────────────
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 22

    ws.freeze_panes = "A6"
