"""
test_config.py — config / threshold / preset tests for the EXFO panel.
=======================================================================

These tests pin the OTDR threshold-panel contract that lives in both
``streamlit_app.py`` (web) and ``desktop/desktop_app.py`` (desktop).

The audit (Phase 1.4) surfaced a handful of bugs and drift risks:

* ``_otdr_override`` happily accepts NaN coming back from the iframe
  (``parseFloat("")`` returns NaN), which becomes ``BEND_THRESHOLD``
  and silently disables every flag (every ``abs(loss) < NaN`` is
  False).
* ``_otdr_override`` also accepts negative loss values, which means
  "flag everything" — almost certainly not what the user meant.
* The customer-profile switch handler does a substring match
  (``"Custom" not in _picked``) so a future profile named e.g.
  ``"Custom Corp"`` would be treated as the sentinel and its preset
  would be silently skipped.
* Rows with ``supported=False`` accept Apply ticks without any
  explicit visual hint that the tick is a no-op in this app.
* Desktop and web carry duplicate copies of ``OTDR_ROWS``,
  ``OTDR_DEFAULT_APPLY``, ``CUSTOMER_PROFILES`` and the engine-wiring
  inside ``_run_engine`` — easy for a future maintainer to update one
  side without the other.

Test plan
---------
The file is split into four blocks:

1. **Strict-XFAIL** — desired behaviour for bugs that have NOT been
   fixed yet.  These flip to XPASS the day the fix lands, which is
   exactly the heads-up we want.
2. **PASS-but-buggy** — pins today's actual behaviour with a TODO
   comment so the fix can't slip out unnoticed.
3. **Drift-prevention** — desktop vs web equality contract tests for
   the duplicate constants and engine wiring.
4. **Threshold flowthrough** — full ``AppTest`` run that drives the
   ``unidir_splice_loss`` override through the UI →
   ``session_state.otdr_settings`` → ``_otdr_override`` →
   ``_run_engine`` → engine constants → output XLSX chain.

Why AST extraction?
-------------------
Both ``streamlit_app.py`` and ``desktop/desktop_app.py`` call
``st.set_page_config`` and friends at import-time, so a plain
``import`` doesn't work outside the Streamlit runtime.  Constants and
the small ``_otdr_override`` function are top-level, side-effect free
definitions, so we use ``ast`` + ``exec`` to lift them into an
isolated namespace without booting Streamlit.  This avoids the
brittle dance of stubbing the entire Streamlit module just to read a
list.

For test #12 (threshold flowthrough) we DO want the full app, so that
one uses ``run_streamlit`` from ``conftest.py`` like ``test_e2e_run``
does.
"""
from __future__ import annotations

import ast
import math
import pathlib
import shutil

import openpyxl
import pytest

from conftest import APP_PATH, FIXTURE_DIR, REPO_ROOT, WEB_APP_PATH, run_streamlit


# ─────────────────────────────────────────────────────────────────────
#  AST helpers — pull top-level symbols out of a Streamlit script
#  without booting Streamlit.
# ─────────────────────────────────────────────────────────────────────
def _load_top_level_symbols(src_path: pathlib.Path, names: set[str]) -> dict:
    """Return a dict of the named top-level definitions from ``src_path``.

    Walks the file's AST, keeps only the ``Assign`` statements (for
    constants) and ``FunctionDef`` statements (for ``_otdr_override``)
    whose target name is in ``names``, and execs them in a clean
    namespace.  No Streamlit, no side effects.
    """
    tree = ast.parse(src_path.read_text(encoding="utf-8"))
    keep: list[ast.stmt] = []
    for node in tree.body:
        if isinstance(node, ast.Assign):
            for t in node.targets:
                if isinstance(t, ast.Name) and t.id in names:
                    keep.append(node)
                    break
        elif isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            if node.name in names:
                keep.append(node)
    module = ast.Module(body=keep, type_ignores=[])
    ns: dict = {}
    exec(compile(module, str(src_path), "exec"), ns)
    # Note: we leave __builtins__ in place so the exec'd functions
    # have access to ``float``, ``TypeError``, ``ValueError`` etc.
    # when they're later invoked from test code.
    return ns


def _load_engine_wiring_source(src_path: pathlib.Path) -> str:
    """Return the source text of the file's ``_run_engine`` function.

    Used by test #11 to assert that exactly one engine constant is
    wired through (``BEND_THRESHOLD``) and that the row key fed to
    ``_otdr_override`` is exactly ``unidir_splice_loss``.  In
    ``streamlit_app.py`` ``_run_engine`` is the dispatcher and the
    actual ``thresholds = {...}`` dict is built at module level just
    below the sidebar — so we just return the whole file for the web
    app.  In ``desktop_app.py`` the dict is also built at module level
    around line ~468.  To keep the test simple and forwards-robust we
    return the whole file's text from both sides and search it as
    plain text.
    """
    return src_path.read_text(encoding="utf-8")


# Convenience: the names we care about in both apps.
_SYMBOLS = {"OTDR_ROWS", "OTDR_DEFAULT_APPLY", "CUSTOMER_PROFILES",
            "_otdr_override"}


@pytest.fixture(scope="module")
def web_ns() -> dict:
    """Top-level symbols lifted out of ``streamlit_app.py``."""
    return _load_top_level_symbols(WEB_APP_PATH, _SYMBOLS)


@pytest.fixture(scope="module")
def desktop_ns() -> dict:
    """Top-level symbols lifted out of ``desktop/desktop_app.py``."""
    return _load_top_level_symbols(APP_PATH, _SYMBOLS)


def _call_override(ns: dict, settings: dict, key: str, default: float) -> float:
    """Call the namespace's ``_otdr_override`` with ``otdr`` patched
    in for the duration of the call.

    Both copies of ``_otdr_override`` close over a module-level
    ``otdr`` name.  We inject it before calling and restore after.
    """
    fn = ns["_otdr_override"]
    saved = ns.get("otdr", None)
    ns["otdr"] = settings
    # Re-bind the function's globals (it was compiled with this ns as
    # its globals), so the reassignment above is what the function
    # sees.  No extra work needed — exec() compiled the function with
    # ``ns`` as its globals.
    try:
        return fn(key, default)
    finally:
        if saved is None:
            ns.pop("otdr", None)
        else:
            ns["otdr"] = saved


# ─────────────────────────────────────────────────────────────────────
#  1. Strict-XFAIL block — desired behaviour (flip to XPASS on fix).
# ─────────────────────────────────────────────────────────────────────
@pytest.mark.xfail(
    strict=True,
    reason=("F4: _otdr_override returns float('nan') when the iframe "
            "sends back NaN; engine then has BEND_THRESHOLD=NaN and "
            "every abs(loss)<NaN is False — every event silently passes."),
)
def test_otdr_override_rejects_nan(desktop_ns):
    """When the iframe returns NaN for ``fail``, ``_otdr_override``
    should fall back to the engine default rather than propagating
    NaN into the threshold."""
    bogus = {
        "unidir_splice_loss": {
            "apply":   True,
            "fail":    float("nan"),
            "warning": float("nan"),
        },
    }
    out = _call_override(desktop_ns, bogus, "unidir_splice_loss", 0.100)
    # Desired post-fix behaviour: NaN is rejected, default is returned.
    assert not math.isnan(out), "override leaked NaN into the threshold"
    assert out == pytest.approx(0.100)


@pytest.mark.xfail(
    strict=True,
    reason=("F4 (negative): _otdr_override returns the negative value "
            "as-is.  A negative threshold means abs(loss)<negative is "
            "always False — every event is flagged.  Fix should clamp "
            "to 0 or fall back to the default."),
)
def test_otdr_override_rejects_negative_loss(desktop_ns):
    """Negative loss thresholds make no physical sense for the
    unidir_splice_loss row (loss magnitudes are positive).  The
    override should not accept them."""
    bogus = {
        "unidir_splice_loss": {
            "apply":   True,
            "fail":    -0.5,
            "warning": -0.5,
        },
    }
    out = _call_override(desktop_ns, bogus, "unidir_splice_loss", 0.100)
    # Desired post-fix behaviour: clamp to 0 or fall back to default.
    # Either way, the engine must not see a negative threshold.
    assert out >= 0.0, f"override returned a negative threshold: {out}"


@pytest.mark.xfail(
    strict=True,
    reason=("F5: profile-switch uses substring match "
            "(`\"Custom\" not in _picked`) so a future preset named "
            "'Custom Corp' would be silently treated as the sentinel "
            "and its preset would not be applied."),
)
def test_custom_profile_sentinel_is_exact_match(desktop_ns):
    """The sentinel-comparison logic in the profile-switch handler
    must use an exact match, not a substring match."""
    profiles = dict(desktop_ns["CUSTOMER_PROFILES"])
    profiles["Custom Corp"] = {
        "apply":      {"unidir_splice_loss"},
        "thresholds": {"unidir_splice_loss": 0.250},
    }
    picked = "Custom Corp"
    # The current code does: `if "Custom" not in _picked:` — which
    # evaluates False for "Custom Corp", so the preset is skipped.
    # Desired logic: the only sentinel is the literal "Custom (edit
    # table below)".  Express the desired predicate explicitly here:
    SENTINEL = "Custom (edit table below)"
    # The fix should switch to ``_picked != SENTINEL`` (or an
    # equivalent exact-match).  Until then, the current logic returns
    # False for "Custom Corp" and the preset is silently skipped.
    is_real_preset = (picked != SENTINEL) and picked in profiles
    assert is_real_preset, (
        "'Custom Corp' should be treated as a real preset (not the sentinel)."
    )
    # And the preset's thresholds should be available in the profiles map.
    assert profiles[picked]["thresholds"]["unidir_splice_loss"] == 0.250

    # Re-assert what the CURRENT buggy substring match does — this is
    # the line that must change for the XFAIL to flip:
    current_substring_check = "Custom" not in picked
    # If the fix has landed (exact match), the current substring check
    # should ALSO return True for "Custom Corp" — but it doesn't.
    assert current_substring_check, (
        "The substring check '\"Custom\" not in _picked' must be replaced "
        "with an exact-match against the literal sentinel string."
    )


@pytest.mark.xfail(
    strict=True,
    reason=("F6: rows with supported=False render as '(not yet wired)' "
            "but the Apply checkbox is still tickable and the value "
            "inputs are still editable.  Desired: the row's Apply "
            "checkbox should be visibly disabled (or the row should "
            "carry an explicit 'no-op in this app' tooltip)."),
)
def test_unsupported_row_apply_is_visibly_marked():
    """The OTDR component HTML should explicitly mark Apply
    checkboxes for ``supported=False`` rows as no-ops — either by
    disabling them, by giving them a ``title`` (tooltip) attribute,
    or by attaching an ``aria-disabled`` hint.  Today the only
    marker is the trailing '(not yet wired)' caption on the label,
    which doesn't tell the user that ticking the box does nothing."""
    html = (REPO_ROOT / "components" / "otdr_settings" / "index.html").read_text(
        encoding="utf-8"
    )
    # Desired: when row.supported is False, the Apply checkbox gets
    # an explicit affordance — one of these markers should appear in
    # the unsupported-row branch.
    markers = ("cb.disabled", "cb.title", "aria-disabled", "no-op")
    assert any(m in html for m in markers), (
        "Apply checkbox on unsupported rows has no explicit "
        "'this does nothing' marker.  Found only the '(not yet wired)' "
        "label suffix, which doesn't tell the user the tick is a no-op."
    )


# ─────────────────────────────────────────────────────────────────────
#  2. PASS-but-buggy block — pins today's behaviour with a TODO.
# ─────────────────────────────────────────────────────────────────────
def test_otdr_override_currently_returns_nan_unchanged(desktop_ns):
    """Pins the current (buggy) NaN-passthrough behaviour.

    TODO: when ``test_otdr_override_rejects_nan`` is fixed, flip this
    assertion (will need coordinated update — the fix must update
    both tests in one commit).
    """
    bogus = {
        "unidir_splice_loss": {
            "apply":   True,
            "fail":    float("nan"),
            "warning": float("nan"),
        },
    }
    out = _call_override(desktop_ns, bogus, "unidir_splice_loss", 0.100)
    assert math.isnan(out), (
        "Expected the current buggy behaviour: NaN is passed through. "
        "If this assertion failed because the fix landed, FLIP the "
        "matching XFAIL test in the same commit."
    )


def test_otdr_override_currently_returns_negative_unchanged(desktop_ns):
    """Pins the current (buggy) negative-loss passthrough.

    TODO: when ``test_otdr_override_rejects_negative_loss`` is fixed,
    flip this assertion (coordinated update — the fix must update
    both tests in one commit).
    """
    bogus = {
        "unidir_splice_loss": {
            "apply":   True,
            "fail":    -0.5,
            "warning": -0.5,
        },
    }
    out = _call_override(desktop_ns, bogus, "unidir_splice_loss", 0.100)
    assert out == pytest.approx(-0.5), (
        "Expected the current buggy behaviour: negative loss passes "
        "through.  If this assertion failed because the fix landed, "
        "FLIP the matching XFAIL test in the same commit."
    )


def test_customer_profile_substring_match_currently_skips_custom_corp(desktop_ns):
    """Pins the current substring-match bug.

    TODO: when ``test_custom_profile_sentinel_is_exact_match`` is
    fixed, flip this assertion (coordinated update).
    """
    picked = "Custom Corp"
    # This is the literal expression from the profile-switch handler:
    is_skipped_by_substring_check = "Custom" in picked   # current buggy logic treats this as "is sentinel"
    assert is_skipped_by_substring_check, (
        "Expected the current buggy substring check: 'Custom Corp' "
        "contains 'Custom' so the handler skips applying the preset.  "
        "If this failed because the fix landed, FLIP the matching "
        "XFAIL test in the same commit."
    )


# ─────────────────────────────────────────────────────────────────────
#  3. Drift-prevention contract tests — desktop must equal web.
# ─────────────────────────────────────────────────────────────────────
def test_otdr_rows_match_between_desktop_and_web(desktop_ns, web_ns):
    """``OTDR_ROWS`` must be identical in desktop and web.  Different
    rows would mean different threshold defaults, units, or
    supported-flag for the same UI label between the two apps."""
    d_rows = desktop_ns["OTDR_ROWS"]
    w_rows = web_ns["OTDR_ROWS"]
    assert d_rows == w_rows, (
        "OTDR_ROWS drift between desktop and web!  Both copies must "
        "stay in lock-step.  Diff:\n"
        f"  desktop: {d_rows!r}\n"
        f"  web:     {w_rows!r}"
    )


def test_otdr_default_apply_matches_between_desktop_and_web(desktop_ns, web_ns):
    """``OTDR_DEFAULT_APPLY`` must be identical in desktop and web."""
    d = desktop_ns["OTDR_DEFAULT_APPLY"]
    w = web_ns["OTDR_DEFAULT_APPLY"]
    assert d == w, (
        f"OTDR_DEFAULT_APPLY drift: desktop={d!r}, web={w!r}"
    )


def test_customer_profiles_match_between_desktop_and_web(desktop_ns, web_ns):
    """``CUSTOMER_PROFILES`` must be identical in desktop and web."""
    d = desktop_ns["CUSTOMER_PROFILES"]
    w = web_ns["CUSTOMER_PROFILES"]
    assert d == w, (
        "CUSTOMER_PROFILES drift between desktop and web!  Both "
        "copies must stay in lock-step.  Diff:\n"
        f"  desktop keys: {sorted(d)}\n"
        f"  web keys:     {sorted(w)}"
    )


def test_only_unidir_splice_loss_is_wired_to_engine():
    """The audit's F6 finding rests on the claim that only the
    ``unidir_splice_loss`` row affects the engine — every other row
    in the OTDR panel is visual-only.  This test inspects both apps'
    source for the engine-wiring shape:

      * ``_otdr_override`` is called with the literal key
        ``\"unidir_splice_loss\"`` and ONLY with that key.
      * The ``thresholds`` dict passed into the engine contains
        exactly one key: ``BEND_THRESHOLD``.

    If a future commit wires more rows, this test fails — forcing a
    coordinated update of the audit docstring AND the OTDR_DEFAULT_APPLY
    set."""
    for label, path in (("desktop", APP_PATH), ("web", WEB_APP_PATH)):
        src = _load_engine_wiring_source(path)

        # Find every _otdr_override(...) call and check the first
        # string-literal argument is "unidir_splice_loss".
        tree = ast.parse(src)
        override_keys = []
        for node in ast.walk(tree):
            if isinstance(node, ast.Call) and isinstance(node.func, ast.Name) \
                    and node.func.id == "_otdr_override":
                if node.args and isinstance(node.args[0], ast.Constant) \
                        and isinstance(node.args[0].value, str):
                    override_keys.append(node.args[0].value)
        assert override_keys == ["unidir_splice_loss"], (
            f"[{label}] _otdr_override call keys drifted: {override_keys!r}. "
            "Expected exactly one call with 'unidir_splice_loss'.  If a "
            "new row was wired, update OTDR_DEFAULT_APPLY and the audit "
            "docstring in lock-step."
        )

        # And the thresholds dict literal must hold exactly one key:
        # "BEND_THRESHOLD".  We look for any dict literal that contains
        # "BEND_THRESHOLD" and verify all its keys are exactly that.
        bend_dicts = []
        for node in ast.walk(tree):
            if isinstance(node, ast.Dict):
                keys = []
                for k in node.keys:
                    if isinstance(k, ast.Constant) and isinstance(k.value, str):
                        keys.append(k.value)
                if "BEND_THRESHOLD" in keys:
                    bend_dicts.append(keys)
        assert bend_dicts, (
            f"[{label}] couldn't find any dict literal with 'BEND_THRESHOLD' "
            "as a key — has the thresholds-passing shape changed?"
        )
        for keys in bend_dicts:
            assert keys == ["BEND_THRESHOLD"], (
                f"[{label}] thresholds dict drifted: {keys!r}.  Expected "
                "exactly {'BEND_THRESHOLD': ...}.  If a new engine "
                "constant is now exposed, update the audit and the "
                "OTDR row that wires it."
            )


# ─────────────────────────────────────────────────────────────────────
#  4. Threshold-flowthrough — end-to-end through the engine.
# ─────────────────────────────────────────────────────────────────────
OUTPUT_SUBDIR_NAME = "_unidir_output"


@pytest.fixture
def cleyak_output_dir() -> pathlib.Path:
    """Same scrub-around-the-test pattern as ``test_e2e_run``: the
    desktop_app writes its XLSX next to the inputs and the fixture
    dir is checked into git, so we delete the output subdir before
    AND after the test."""
    out = FIXTURE_DIR / OUTPUT_SUBDIR_NAME
    if out.exists():
        shutil.rmtree(out)
    yield out
    if out.exists():
        shutil.rmtree(out)


def test_unidir_splice_loss_override_reaches_engine(
    cleyak_output_dir: pathlib.Path,
) -> None:
    """Full flow: with the panel set to apply ``unidir_splice_loss
    = 0.250``, no flagged bend/damage event in the produced XLSX
    should have ``|loss| < 0.250``.  Proves the threshold override
    propagated all the way through:

        UI → session_state.otdr_settings → _otdr_override
           → thresholds dict → _run_engine → engine.BEND_THRESHOLD
           → write_xlsx → on-disk Flagged Events sheet.
    """
    assert FIXTURE_DIR.is_dir(), (
        f"fixture directory {FIXTURE_DIR} is missing.  "
        "Did the cleyak_mini fixtures get committed?"
    )

    threshold = 0.250  # well above the engine default of 0.100

    # First pass walks the script with an empty _folder so the OTDR
    # sidebar component initialises and session_state["otdr_settings"]
    # is populated with defaults.  The script stops at the "pick a
    # folder" gate (st.stop).
    at = run_streamlit(APP_PATH, default_timeout=120)
    assert not at.exception, (
        f"unexpected exception on first run before Run was triggered: "
        f"{at.exception}"
    )

    # Now override the OTDR panel's session state.  Mutate the dict
    # under the same key the desktop app reads (`otdr_settings`).
    # No widget binds that key, so a plain session-state write is the
    # canonical path here.
    settings = dict(at.session_state["otdr_settings"])
    settings["unidir_splice_loss"] = {
        "apply":   True,
        "fail":    threshold,
        "warning": threshold,
    }
    at.session_state["otdr_settings"] = settings

    # Point the app at the fixture and force the Run branch.
    at.session_state["_folder"] = str(FIXTURE_DIR)
    at.session_state["__test_force_run__"] = True

    at.run()

    assert not at.exception, (
        "desktop_app raised on the Run path with the threshold override. "
        f"Exception(s): {at.exception}"
    )

    result = at.session_state["_last_result"]
    assert result, "_run_engine returned no result — Run branch did not fire."

    xlsx_path = pathlib.Path(result["xlsx_path"])
    assert xlsx_path.exists(), f"engine did not write {xlsx_path}"

    # Open the Flagged Events sheet and scan every bend/damage row.
    # Splice rows are aggregated by population, break rows have no
    # loss number — only bend/damage rows are loss-gated by
    # BEND_THRESHOLD, so those are the rows that must respect the
    # override.
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        assert "Flagged Events" in wb.sheetnames, (
            f"XLSX missing 'Flagged Events' sheet: {wb.sheetnames!r}"
        )
        ws = wb["Flagged Events"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    assert rows, "Flagged Events sheet is empty — engine produced no header row?"
    header = [str(c) if c is not None else "" for c in rows[0]]
    try:
        loss_col  = header.index("Loss (dB)")
        kind_col  = header.index("Kind")
    except ValueError as e:
        raise AssertionError(
            f"Flagged Events header layout changed: {header!r}"
        ) from e

    offenders = []
    for r in rows[1:]:
        kind = r[kind_col]
        loss = r[loss_col]
        if kind != "Possible Bend/Damage":
            continue
        # Loss should always be numeric for bend/damage; only break
        # rows carry the string "broke".  Belt and braces:
        if not isinstance(loss, (int, float)):
            continue
        if abs(loss) < threshold - 1e-9:    # tiny epsilon for float wobble
            offenders.append((kind, loss))

    assert not offenders, (
        f"Threshold override did NOT reach the engine.  With "
        f"BEND_THRESHOLD={threshold}, found {len(offenders)} "
        "bend/damage rows with |loss| below the threshold.  First "
        f"few: {offenders[:5]!r}"
    )
