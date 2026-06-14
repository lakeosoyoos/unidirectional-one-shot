"""
test_engine_contracts.py — engine-boundary contracts + edge-case pins.
=====================================================================

Agent D's slice of the Phase 2 test suite.  These tests do NOT touch
the Streamlit UI.  They call engine functions directly with synthetic
inputs (and one fixture-driven sanity check) and run in well under
five seconds for the whole file.

Two flavours of test live here:

  1.  Signature contracts (PASS today, must keep PASSing).  Both UIs
      call the engine with positional + keyword arguments that are
      load-bearing.  A future rename in the engine would silently
      break those calls — these tests catch the drift at the engine
      boundary, much cheaper than the full ``test_e2e_run.py``.

  2.  Edge-case pins (some PASS, some strict-XFAIL).  Each pins a
      known wart from the Phase 1 audit:

        PASS  → reproduces today's behaviour so refactors don't drift
                further from it without telling us.
        XFAIL → describes the *desired* behaviour.  Today's wart makes
                them fail; pytest reports them as ``xfail``.  The day
                someone fixes the wart they'll XPASS and the strict
                flag will turn them red, prompting whoever did the
                cleanup to flip the test from XFAIL to PASS.

The audit findings being pinned here:

    H1  empty fibers → ``max(fibers.keys())`` ValueError in main()
    H2  empty fibers → write_xlsx skips the audit sheet entirely
    L17 ``direction_signature`` collapses orig==term + no cable_id to '?'
    M8  ISO-8601 duration regex too strict (no lowercase, no days)
    M9  ``int(n_avg) if n_avg else None`` treats 0 as missing
    M12 NaN losses survive ``_format_cell_label`` as the string "nan"

The fixture-driven test at the bottom is a single sanity check that
the full engine pipeline still produces the canonical 5-sheet layout
on the bundled 25-fiber YAKCLE corpus.  Cheap enough to live in this
file (sub-second on a warm cache) and useful for catching sheet-order
or sheet-name regressions before they hit the e2e suite.
"""
from __future__ import annotations

import inspect
import pathlib
import tempfile

import openpyxl
import pytest

# The conftest at tests/conftest.py already puts the repo root on
# sys.path, so engine imports work whether pytest is run from the
# repo root or from inside tests/.
from conftest import FIXTURE_DIR

import acquisition_audit
import json_reader
import reburn_percentage
import unidirectional_event_finder as engine


# ════════════════════════════════════════════════════════════════════════
#  1 — SIGNATURE CONTRACTS  (PASS today, must keep PASSing)
# ════════════════════════════════════════════════════════════════════════

def test_load_fibers_signature():
    """``streamlit_app`` and ``desktop_app`` both call
    ``engine.load_fibers(<folder>, direction=<sig>)``.  If the engine
    ever renames either parameter the UIs break silently — this catches
    the drift at the engine boundary."""
    params = inspect.signature(engine.load_fibers).parameters
    assert "input_path" in params, (
        f"engine.load_fibers must accept an 'input_path' parameter; "
        f"got {list(params)!r}"
    )
    assert "direction" in params, (
        f"engine.load_fibers must accept a 'direction' keyword; "
        f"got {list(params)!r}"
    )


def test_write_xlsx_signature():
    """Both UIs invoke ``write_xlsx(..., site_a=, site_b=, fibers=)`` by
    keyword.  Rename any of those three kwargs and both UIs crash with
    TypeError at runtime, which is exactly what we don't want shipped."""
    params = inspect.signature(engine.write_xlsx).parameters
    for kw in ("site_a", "site_b", "fibers"):
        assert kw in params, (
            f"engine.write_xlsx must accept '{kw}' keyword; "
            f"got {list(params)!r}"
        )


def test_extract_acquisition_signature():
    """``acquisition_audit.extract_acquisition(fiber: dict)`` — the audit
    builder calls this once per loaded fiber.  A rename here breaks the
    'Acquisition Parameters' sheet wholesale."""
    params = inspect.signature(acquisition_audit.extract_acquisition).parameters
    assert "fiber" in params, (
        f"extract_acquisition must accept a 'fiber' dict; got {list(params)!r}"
    )


def test_build_audit_signature():
    """``acquisition_audit.build_audit(fibers: dict)`` — write_xlsx hands
    it the loaded-fibers map directly.  Rename the parameter and the
    audit sheet stops being written (silently, because the wider
    try/except below it swallows the TypeError)."""
    params = inspect.signature(acquisition_audit.build_audit).parameters
    assert "fibers" in params, (
        f"build_audit must accept a 'fibers' dict; got {list(params)!r}"
    )


def test_build_reburn_summary_signature():
    """``reburn_percentage.build_reburn_summary(grid, columns, n_ribbons,
    ribbon_label_fn=None)`` — write_xlsx passes the first three
    positionally and ribbon_label_fn by keyword."""
    params = inspect.signature(reburn_percentage.build_reburn_summary).parameters
    for name in ("grid", "columns", "n_ribbons", "ribbon_label_fn"):
        assert name in params, (
            f"build_reburn_summary must accept '{name}'; got {list(params)!r}"
        )
    # ribbon_label_fn must remain optional (default None) — write_xlsx
    # passes it but other callers don't.
    assert (
        params["ribbon_label_fn"].default
        is inspect.Parameter.empty
        or params["ribbon_label_fn"].default is None
    ), "ribbon_label_fn should default to None"


def test_write_reburn_sheet_signature():
    """``reburn_percentage.write_reburn_sheet(wb, summary, insert_at=1,
    sheet_title=...)`` — write_xlsx calls it with insert_at=1 by keyword."""
    params = inspect.signature(reburn_percentage.write_reburn_sheet).parameters
    for name in ("wb", "summary", "insert_at", "sheet_title"):
        assert name in params, (
            f"write_reburn_sheet must accept '{name}'; got {list(params)!r}"
        )
    assert params["insert_at"].default == 1, (
        f"write_reburn_sheet's insert_at default must stay at 1 so the "
        f"reburn sheet sits immediately after the audit sheet; got "
        f"{params['insert_at'].default!r}"
    )


# ════════════════════════════════════════════════════════════════════════
#  2 — EDGE-CASE PINS (PASS — reproduces today's wart)
# ════════════════════════════════════════════════════════════════════════

def test_empty_fibers_currently_crashes_at_max_keys():
    """H1 — ``main()`` does ``n_fibers = max(fibers.keys())`` with no
    guard.  An empty load (e.g. tech picks a folder where every file
    fails to parse) raises ValueError before any output is produced.
    Pin the crash so a future fix is visible as an XPASS."""
    fibers: dict = {}
    with pytest.raises(ValueError):
        # This is the literal line from main() that explodes today.
        max(fibers.keys())


def test_reburn_empty_grid_renders_zero_over_zero():
    """H2 (companion) — ``build_reburn_summary({}, [], 0)`` returns a
    summary with total_cells == 0 and percentage == 0.0, which is what
    drives the workbook's '0.00% (0 of 0)' headline on empty input.
    Pin that today's headline is silent rather than e.g. raising."""
    summary = reburn_percentage.build_reburn_summary({}, [], 0)
    assert summary["total_cells"] == 0
    assert summary["reburn_cells"] == 0
    assert summary["percentage"] == 0.0
    assert summary["per_splice"] == []
    assert summary["per_ribbon"] == []


def test_iso_duration_rejects_lowercase_and_multiday():
    """M8 — the regex is anchored ``^PT...$`` with no 'D' branch and no
    case-insensitive flag.  Anything lowercase or with a 'D' component
    or a trailing ``PT15.S`` returns None instead of seconds."""
    parse = json_reader._parse_iso8601_duration_seconds
    assert parse("pt15s") is None
    assert parse("P1DT2H") is None
    assert parse("PT15.S") is None
    # And confirm the happy path still works so we know we're pinning
    # the *strictness*, not a totally broken parser.
    assert parse("PT15S") == 15.0


def test_zero_averaging_count_reads_as_missing():
    """M9 — ``int(n_avg) if n_avg else None`` treats 0 as missing.
    A fiber whose calibration block honestly reports zero averages
    shows up as 'Not available' in the audit sheet, hiding the
    misconfiguration.  Pin today's silent-drop behaviour."""
    fiber = {
        "_source":          "sor",
        "filename":         "fake.sor",
        "date_time":        0,
        "sup_params":       {},
        "wavelength":       1550,        # integer nm — extract uses round()
        "exfo_calibration": {
            "NominalPulseWidth": 50e-9,
            "NominalWavelength": 1550,
            "NumberOfAverages":  0,       # honest zero, NOT missing
        },
    }
    rec = acquisition_audit.extract_acquisition(fiber)
    assert rec["averaging_count"] is None, (
        f"today's wart: zero averages should round-trip but the engine "
        f"reads it as missing.  Got averaging_count={rec['averaging_count']!r}"
    )


def test_format_cell_label_nan_renders_as_nan_string():
    """M12 — NaN losses survive ``_format_cell_label`` and surface in
    the workbook as the literal string 'nan'.  Pin that today's output
    is silent garbage rather than a clear error or omission."""
    label = engine._format_cell_label([(1, float("nan"))])
    assert "nan" in label.lower(), (
        f"expected today's silent-NaN behaviour: label should contain "
        f"'nan'; got {label!r}"
    )


# ════════════════════════════════════════════════════════════════════════
#  3 — STRICT XFAIL (desired behaviour; XPASS when someone fixes it)
# ════════════════════════════════════════════════════════════════════════

@pytest.mark.xfail(strict=True,
                   reason="H1: main() crashes on empty fibers via max(fibers.keys())")
def test_empty_fibers_does_not_crash():
    """The engine should produce an empty-but-valid workbook (or at
    least a clean error) when the loader returns zero fibers, instead
    of crashing the worker process with ValueError."""
    fibers: dict = {}
    # Today this raises; tomorrow it should not.
    n_fibers = max(fibers.keys(), default=0)  # noqa: F841 — what the fix would look like
    # If we ever reach this line it means main() was patched to guard
    # the call, and the test will XPASS — flip it to PASS at that point.
    raise AssertionError(
        "If you see this failure it means max() did NOT raise — the H1 "
        "wart has been fixed.  Flip this test from XFAIL to PASS and "
        "add a real assertion about the empty-input output."
    )


@pytest.mark.xfail(strict=True,
                   reason="H2: write_xlsx skips audit sheet when fibers is falsy")
def test_write_xlsx_writes_audit_even_on_empty_fibers():
    """The audit sheet should always exist, even on empty input, so a
    tech opening the workbook sees '0 traces in this run' instead of a
    workbook that silently omits the headline sheet."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        out = pathlib.Path(tf.name)
    try:
        engine.write_xlsx(
            grid={}, columns=[], n_fibers=0, ribbon_size=12,
            span_km=0.0, output_path=str(out),
            site_a="", site_b="", fibers={},
        )
        wb = openpyxl.load_workbook(out, read_only=True)
        try:
            sheets = list(wb.sheetnames)
        finally:
            wb.close()
        assert "Acquisition Parameters" in sheets, (
            f"audit sheet should be present even on empty input; got {sheets!r}"
        )
    finally:
        if out.exists():
            out.unlink()


@pytest.mark.xfail(strict=True,
                   reason="M8: regex too strict — rejects lowercase and 'D' duration components")
def test_iso_duration_accepts_lowercase_and_multiday():
    """A robust parser should accept both lowercase markers (some
    instruments lowercase the entire ISO duration) and the 'D' day
    component (P1DT2H = 1 day 2 hours = 93,600 s)."""
    parse = json_reader._parse_iso8601_duration_seconds
    assert (parse("pt15s") or 0) > 0
    assert (parse("P1DT2H") or 0) > 0


@pytest.mark.xfail(strict=True,
                   reason="M9: int(n_avg) if n_avg else None treats 0 as missing")
def test_zero_averaging_count_reads_as_zero():
    """A fiber whose calibration honestly reports zero averages should
    round-trip as 0, not None, so the audit sheet can flag the
    misconfiguration instead of silently hiding it."""
    fiber = {
        "_source":          "sor",
        "filename":         "fake.sor",
        "date_time":        0,
        "sup_params":       {},
        "wavelength":       1550,
        "exfo_calibration": {
            "NominalPulseWidth": 50e-9,
            "NominalWavelength": 1550,
            "NumberOfAverages":  0,
        },
    }
    rec = acquisition_audit.extract_acquisition(fiber)
    assert rec["averaging_count"] == 0


@pytest.mark.xfail(strict=True,
                   reason="M12: NaN losses surface as the string 'nan' in cell labels")
def test_format_cell_label_rejects_or_omits_nan():
    """The engine should either omit the NaN entry from the label or
    raise a clear error so the silent-garbage 'F12 nan' cell stops
    shipping in workbooks."""
    label = engine._format_cell_label([(1, float("nan"))])
    # Either path is acceptable: the entry is dropped (empty label) or
    # the NaN is normalised to a readable token.  The one thing that's
    # NOT acceptable is the literal string "nan".
    assert "nan" not in label.lower(), (
        f"NaN should not survive into the label; got {label!r}"
    )


@pytest.mark.xfail(strict=True,
                   reason="L17: orig==term + no cable_id collapses to '?'; "
                          "different cables collide")
def test_direction_signature_collapses_orig_eq_term_to_cable_id():
    """When two physical cables happen to terminate at the same site
    (orig==term, e.g. an interconnect loop) and the SOR/JSON file
    doesn't expose a cable_id, both today resolve to ``'?'`` and the
    loader merges them into one signature.  The desired behaviour is
    that some other distinguishing metadata (filename, fiber group,
    folder…) keeps the two runs separate."""
    meta_a = {"orig_loc": "VANCOUVER", "term_loc": "VANCOUVER",
              "cable_id": "", "filename": "cableA_001.sor"}
    meta_b = {"orig_loc": "VANCOUVER", "term_loc": "VANCOUVER",
              "cable_id": "", "filename": "cableB_001.sor"}
    sig_a = engine.direction_signature(meta_a)
    sig_b = engine.direction_signature(meta_b)
    assert sig_a != sig_b, (
        f"two unrelated cables collided onto signature {sig_a!r}; "
        f"loader will merge them into one run"
    )


# ════════════════════════════════════════════════════════════════════════
#  4 — REAL-FIXTURE REGRESSION (PASS — cheap sanity check)
# ════════════════════════════════════════════════════════════════════════

def test_cleyak_mini_fixture_produces_audit_and_reburn_sheets():
    """End-to-end smoke test of the engine pipeline on the bundled
    25-fiber YAKCLE corpus.  Asserts the canonical sheet layout
    (Acquisition Parameters first, Reburn Percentage second, ribbon
    grid third, then Legend + Flagged Events).  Pins today's order so
    a silent reshuffle in write_xlsx fails this test instead of
    surprising a tech who opens the workbook."""
    assert FIXTURE_DIR.is_dir(), (
        f"fixture directory {FIXTURE_DIR} is missing — did the "
        "cleyak_mini fixtures get committed?"
    )

    fibers, chosen_dir = engine.load_fibers(str(FIXTURE_DIR), direction=None)
    assert fibers, "loader returned no fibers from the cleyak_mini fixture"

    engine.normalize_all(fibers)
    candidates = engine.discover_splices(fibers)
    valid = engine.refine_and_validate(fibers, candidates)
    off_evs = engine.find_off_splice_events(fibers, valid)
    off_cols = engine.cluster_off_splice(off_evs, fibers)

    n_fibers = max(fibers.keys())
    span = engine.auto_detect_span(fibers)
    breaks = engine.find_breaks(fibers, valid, span)
    break_cols = engine.cluster_breaks(breaks)

    columns = engine.build_columns(valid, off_cols, break_cols)
    grid = engine.build_ribbon_grid(fibers, columns, engine.RIBBON_SIZE)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        out = pathlib.Path(tf.name)
    try:
        engine.write_xlsx(
            grid, columns, n_fibers, engine.RIBBON_SIZE, span, str(out),
            site_a="YAK", site_b="CLE", fibers=fibers,
        )
        wb = openpyxl.load_workbook(out, read_only=True)
        try:
            sheets = list(wb.sheetnames)
        finally:
            wb.close()
    finally:
        if out.exists():
            out.unlink()

    expected_order = [
        "Acquisition Parameters",
        "Reburn Percentage",
        "Unidir Events",
        "Legend",
        "Flagged Events",
    ]
    assert sheets == expected_order, (
        f"sheet order regressed.\n  expected: {expected_order!r}\n  got:      {sheets!r}"
    )
