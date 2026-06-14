"""
test_e2e_run.py — full Run-button flow through the desktop_app.
==================================================================

This is the headline regression test for Phase 2.  It drives the
ENTIRE desktop UI script through ``streamlit.testing.v1.AppTest``:

  1.  Load ``desktop/desktop_app.py``.
  2.  First ``at.run()`` walks the script with an empty ``_folder``
      so the EXFO sidebar component initialises and Step 1 lays out
      its buttons / paste box — then ``st.stop()`` fires at line 504.
  3.  Set ``_folder`` to the bundled fixture directory and set the
      ``__test_force_run__`` session-state hook so the Run button
      branch fires without a real click.
  4.  Second ``at.run()`` walks the script all the way through Step 4
      (inventory → direction detection → engine load → splice
      discovery → XLSX write).
  5.  Assert that no exception was raised AND that the engine
      actually produced an output XLSX with all expected sheets.

The fixture is 25 real YAKCLE JSON files — enough to clear
``MIN_POP_SPLICE = 20`` so the splice-discovery code path actually
executes (not just the empty-result path).

If a future change breaks any link in that chain — folder picker,
inventory, direction signature, engine, XLSX writer — this test
fails BEFORE the Windows build is published.
"""
from __future__ import annotations

import os
import pathlib
import shutil

import openpyxl
import pytest

from conftest import APP_PATH, FIXTURE_DIR, run_streamlit


# Where the desktop_app writes its output when ``_folder`` is a
# directory (vs an extracted .zip): the literal subfolder
# ``_unidir_output`` next to the inputs.
OUTPUT_SUBDIR_NAME = "_unidir_output"


@pytest.fixture
def cleyak_output_dir() -> pathlib.Path:
    """Yield the path the desktop_app will write to, and make sure
    we leave the repo clean afterwards.  The fixture dir is checked
    into git, so we MUST scrub the engine's output between runs."""
    out = FIXTURE_DIR / OUTPUT_SUBDIR_NAME
    # Pre-clean too, in case a previous local run left something behind.
    if out.exists():
        shutil.rmtree(out)
    yield out
    if out.exists():
        shutil.rmtree(out)


def test_run_button_produces_xlsx(cleyak_output_dir: pathlib.Path) -> None:
    """End-to-end: pointing the desktop_app at the cleyak_mini fixture
    and triggering the Run branch must produce a real XLSX with all
    five expected sheets and no script exception."""
    assert FIXTURE_DIR.is_dir(), (
        f"fixture directory {FIXTURE_DIR} is missing.  "
        "Did the cleyak_mini fixtures get committed?"
    )

    at = run_streamlit(APP_PATH, default_timeout=120)

    # First run should have stopped cleanly at the "pick a folder" gate.
    assert at.exception == [] or not at.exception, (
        f"unexpected exception on first run before Run was triggered: "
        f"{at.exception}"
    )

    # Simulate the tech having picked the fixture folder + clicking Run.
    at.session_state["_folder"] = str(FIXTURE_DIR)
    at.session_state["__test_force_run__"] = True

    at.run()

    # The Streamlit `at.exception` ElementList is empty on success.
    assert not at.exception, (
        "desktop_app raised on the Run path with the cleyak_mini fixture. "
        f"Exception(s): {at.exception}"
    )

    # _run_engine stores its return dict under this session key.
    result = at.session_state["_last_result"]
    assert result, "_run_engine returned no result — Run branch did not fire."

    # Engine sanity — the 25-fiber slice should load cleanly.
    assert result["n_fibers"] == 25, (
        f"expected 25 fibers from the cleyak_mini fixture, got "
        f"{result['n_fibers']}"
    )

    # XLSX is actually on disk where the UI claims it is.
    xlsx_path = pathlib.Path(result["xlsx_path"])
    assert xlsx_path.exists(), f"engine did not write {xlsx_path}"
    assert xlsx_path.stat().st_size > 0, f"{xlsx_path} is empty"

    # And the workbook opens with the five expected sheets.
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        sheetnames = list(wb.sheetnames)
    finally:
        wb.close()

    assert len(sheetnames) >= 1, (
        f"XLSX has no sheets at all: {sheetnames!r}"
    )
    # The known canonical layout (Acquisition Parameters first,
    # Reburn Percentage second, then the data sheets).  We assert each
    # by name so a silent rename also fails this test.
    expected_sheets = {
        "Acquisition Parameters",
        "Reburn Percentage",
        "Unidir Events",
        "Legend",
        "Flagged Events",
    }
    missing = expected_sheets - set(sheetnames)
    assert not missing, (
        f"XLSX is missing expected sheets: {missing}.  "
        f"Got: {sheetnames!r}"
    )


def test_fixture_dir_is_under_size_budget() -> None:
    """The fixture corpus must stay tiny so the test suite is cheap
    to run in CI and the repo size doesn't balloon.  10 MB hard cap."""
    total = 0
    for dirpath, _, files in os.walk(FIXTURE_DIR):
        for fn in files:
            total += (pathlib.Path(dirpath) / fn).stat().st_size
    mb = total / (1024 * 1024)
    assert mb <= 10.0, (
        f"tests/fixtures/cleyak_mini is {mb:.2f} MB — over the 10 MB cap. "
        "Prune fixtures or split into a separate downloadable corpus."
    )
