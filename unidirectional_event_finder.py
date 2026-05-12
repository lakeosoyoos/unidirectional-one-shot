#!/usr/bin/env python3
"""
unidirectional_event_finder.py
==============================

A-direction-only event finder.

Goal:
  1. Discover splice closure positions from the A-side population.
  2. Find every other A-side event that lives off-splice.
  3. Render a ribbon-grid Excel where each column (splice or off-splice
     cluster) is shaded for every ribbon that has at least one fiber
     with an event at that distance.

Inputs accepted (auto-detected, single positional argument):
  - Directory of .sor files
  - Directory of .json files (EXFO FastReporter exports)
  - Directory containing .zip files of either / both
  - A single .zip file containing either / both
  - A directory mixing loose files and zips

Closure-discovery math is identical to the bidirectional splice report
(splicereportmatchexfo_April23_AOrdered.py) — 1 km population bins with
MIN_POP_SPLICE fibers, mode-peak refinement, loss-distribution
phantom-rejection.
"""
from __future__ import annotations

import argparse
import json
import math
import os
import struct
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict

import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

from sor_reader324802a import parse_sor_full
from json_reader import parse_otdr_json


# ── Constants (mirrored from splicereportmatchexfo_April23_AOrdered.py) ──
RIBBON_SIZE                   = 12
MIN_POP_SPLICE                = 20      # min fibers to declare a candidate closure
LAUNCH_FIBER_MAX              = 3.0     # km — launch-end exclusion
END_REGION_KM                 = 0.5     # km — drop events this close to fiber end
BEND_THRESHOLD                = 0.100   # dB — universal unidirectional flag threshold
                                        #      (loss below this is not flagged anywhere —
                                        #      splice columns AND off-splice).  Per tech
                                        #      direction May 2026: the uni one-shot only
                                        #      cares about events ≥ 0.1 dB.
CLOSURE_MATCH_KM              = 0.150   # km — events within this of a valid closure are at-splice
CLOSURE_MODE_BIN_M            = 25
CLOSURE_MODE_WINDOW_M         = 75
CLOSURE_VALID_MIN_GAINER_FRAC = 0.05
CLOSURE_VALID_MEDIAN_LOSS_MAX = 0.100
OFF_SPLICE_CLUSTER_M          = 100     # m — merge off-splice events within this distance
HIGH_LOSS_STANDALONE_DB       = 0.100   # dB — same as BEND_THRESHOLD now; kept for
                                        #      readability where it tags "≥0.1 dB" columns.
BREAK_PREMATURE_KM            = 3.0     # km — fiber EOF must be at least this far short
                                        #      of the cable span to be flagged as a break


# ═══════════════════════════════════════════════════════════════════════
#  STEP 1 — Collect input files (handles directories + zips)
# ═══════════════════════════════════════════════════════════════════════

def _is_supported(name: str) -> bool:
    n = name.lower()
    return n.endswith('.sor') or n.endswith('.json')


def _extract_fiber_num(filename: str):
    """Extract the fiber number from an OTDR filename, robustly.

    Pure filename heuristic — only used when GenParams metadata doesn't
    provide a fiber_id.  Strategy:

      1. Split basename by underscore (and trailing whitespace).  For
         each chunk, take its trailing digit run and return that value
         if it's in the plausible fiber range (1–9999).  This handles
         names like ``YAKCLE001_1550`` → "001" → 1 (the "1550" chunk
         exceeds 9999 only if you treat 1550 as a fiber, which we
         intentionally accept; but the underscore-separated "001"
         chunk is checked first).

      2. Fall back to all digit runs in the basename, preferring
         zero-padded runs (``0001`` → 1) over un-padded ones.

      3. Final fallback: take the last 4 digits of the trailing run
         (handles ``CLE1CLE20001`` → trailing run "20001" → last 4
         "0001" → 1).
    """
    base = os.path.basename(filename).rsplit('.', 1)[0].strip()

    # Strategy 1: underscore-separated chunks
    for part in base.split('_'):
        part = part.strip()
        digits = ''
        for ch in reversed(part):
            if ch.isdigit():
                digits = ch + digits
            else:
                break
        if digits:
            n = int(digits)
            if 1 <= n <= 9999:
                return n

    # Strategy 2: all digit runs, prefer zero-padded
    runs = []
    cur = ''
    for ch in base:
        if ch.isdigit():
            cur += ch
        else:
            if cur:
                runs.append(cur)
                cur = ''
    if cur:
        runs.append(cur)
    if not runs:
        return None
    for r in runs:
        if len(r) > 1 and r[0] == '0':
            return int(r)

    # Strategy 3: trailing 4 digits of last run
    last = runs[-1]
    if len(last) > 4:
        last = last[-4:]
    return int(last)


def _walk_files(root: str):
    """Yield (filepath, relname) for every .sor/.json file at root, including
    inside any .zip files (extracted to a temp dir)."""
    if os.path.isfile(root):
        targets = [root]
    else:
        targets = []
        for dirpath, _, fnames in os.walk(root):
            for fn in fnames:
                targets.append(os.path.join(dirpath, fn))

    for path in targets:
        low = path.lower()
        if low.endswith('.zip'):
            try:
                with zipfile.ZipFile(path) as zf:
                    tmpdir = tempfile.mkdtemp(prefix='unidir_')
                    zf.extractall(tmpdir)
                    for sub_dirpath, _, sub_fnames in os.walk(tmpdir):
                        for sub_fn in sub_fnames:
                            if _is_supported(sub_fn):
                                yield os.path.join(sub_dirpath, sub_fn), sub_fn
            except zipfile.BadZipFile:
                print(f"  WARN: bad zip skipped: {path}")
        elif _is_supported(path):
            yield path, os.path.basename(path)


# ── Direction-from-metadata helpers ──────────────────────────────────

def _read_sor_blocks(data: bytes) -> dict:
    """Parse the Bellcore SOR Map (block directory) with proper sequential
    block placement.  Returns {block_name: {offset, body, size}}.

    Differs from the existing sor_reader324802a._parse_block_directory in
    two ways: (1) bounds entry parsing by the Map block's declared total
    size (the existing parser blindly iterates `num_blocks` and can pick
    up a junk 9th entry), and (2) computes block content offsets by
    summing declared sizes, so GenParams (the first block) is found at
    its real location rather than at its directory-entry position.
    """
    name_end = data.index(b'\x00', 0) + 1
    map_size = struct.unpack_from('<I', data, name_end + 2)[0]
    off = name_end + 6
    off += 2                                  # skip num_blocks
    entries = []
    while off < map_size:
        ne = data.index(b'\x00', off) + 1
        nm = data[off:ne - 1].decode('latin-1')
        bv = struct.unpack_from('<H', data, ne)[0]
        bs = struct.unpack_from('<I', data, ne + 2)[0]
        entries.append((nm, bv, bs))
        off = ne + 6
    blocks = {}
    cur = map_size
    for nm, bv, bs in entries:
        body = cur + len(nm) + 1
        blocks[nm] = {'offset': cur, 'body': body, 'size': bs, 'ver': bv}
        cur += bs
    return blocks


def _read_sor_genparams(filepath: str) -> dict:
    """Return the GenParams metadata dict for a SOR file:
        {cable_id, fiber_id, orig_loc, term_loc, cable_code, build_cond,
         operator, comment}
    Any missing field is the empty string.  Returns {} on parse failure."""
    try:
        with open(filepath, 'rb') as f:
            data = f.read()
        blocks = _read_sor_blocks(data)
        if 'GenParams' not in blocks:
            return {}
        body = blocks['GenParams']['body']
        o = body
        # Skip 2-byte language code (e.g. 'EN').
        o += 2

        def pull(o):
            e = data.index(b'\x00', o)
            return data[o:e].decode('latin-1', errors='replace').strip(), e + 1

        cable_id, o   = pull(o)
        fiber_id, o   = pull(o)
        o += 4   # FiberType (2) + NominalWavelength (2)
        orig_loc, o   = pull(o)
        term_loc, o   = pull(o)
        cable_code, o = pull(o)
        # 2-byte BuildCondition
        try:
            build_cond = data[o:o + 2].decode('latin-1', errors='replace')
            o += 2
        except Exception:
            build_cond = ''
        # 8 bytes: UserOffset (4) + UserOffsetDistance (4)
        o += 8
        operator, o = pull(o)
        comment, _  = pull(o)
        return {
            'cable_id': cable_id,
            'fiber_id': fiber_id,
            'orig_loc': orig_loc,
            'term_loc': term_loc,
            'cable_code': cable_code,
            'build_cond': build_cond,
            'operator': operator,
            'comment': comment,
        }
    except Exception:
        return {}


def _read_json_genparams(filepath: str) -> dict:
    """Same shape as _read_sor_genparams, sourced from an EXFO JSON export.

    EXFO FastReporter writes:
        Identification.LocationA / LocationB / OperatorA / Comment / ...
        FiberInformation.LocationDirection = "AB" or "BA"
        Measurement.OtdrMeasurements[0].FiberInformation.LocationDirection
    The LocationDirection flag is what splits the dataset into A→B vs B→A
    runs — same physical fiber, opposite OTDR origin.  We map AB / BA to
    the originating / terminating location pair so direction_signature()
    gets a stable string to group by.
    """
    try:
        with open(filepath) as f:
            obj = json.load(f)
        ident = obj.get('Identification') or {}
        fiber_info = obj.get('FiberInformation') or {}
        m = obj.get('Measurement') or {}
        otdr_list = m.get('OtdrMeasurements') or []
        otdr_fi = (otdr_list[0].get('FiberInformation') or {}) if otdr_list else {}

        def s(v):
            return v.strip() if isinstance(v, str) else ''

        location_a = s(ident.get('LocationA'))
        location_b = s(ident.get('LocationB'))
        location_direction = (
            s(fiber_info.get('LocationDirection'))
            or s(otdr_fi.get('LocationDirection'))
        ).upper()

        # Translate AB/BA into originating/terminating locations
        if location_direction == 'BA' and location_a and location_b:
            orig, term = location_b, location_a
        elif location_direction == 'AB' and location_a and location_b:
            orig, term = location_a, location_b
        else:
            orig, term = location_a, location_b

        # Best-effort fiber id (rarely present, but try)
        fiber_id = ''
        for c in (ident, fiber_info, otdr_fi):
            for k in ('FiberId', 'FiberID', 'Fiber'):
                v = s(c.get(k))
                if v:
                    fiber_id = v
                    break
            if fiber_id:
                break

        return {
            'cable_id':   s(ident.get('JobId')),
            'fiber_id':   fiber_id,
            'orig_loc':   orig,
            'term_loc':   term,
            'cable_code': s(fiber_info.get('FiberType')),
            'build_cond': location_direction,
            'operator':   s(ident.get('OperatorA')),
            'comment':    s(ident.get('Comment')),
        }
    except Exception:
        return {}


def _short_code(location_str: str) -> str:
    """Compress a location string into a 3-letter uppercase code, e.g.
    'Yakima RMU13' → 'YAK', 'Cle Elum' → 'CLE'.  Skips digits and
    non-alphabetic chars.  Returns empty string when nothing usable."""
    if not location_str:
        return ''
    letters = ''.join(c for c in location_str if c.isalpha())[:3]
    return letters.upper()


def direction_signature(meta: dict) -> str:
    """Build a stable string that identifies which direction this trace
    represents.  Prefers an originating→terminating pair when the two
    ends differ; falls back to cable_id (which often encodes direction
    as <fromcode><tocode>, e.g. CLE1CLE2 vs CLE2CLE1)."""
    orig = (meta.get('orig_loc') or '').strip()
    term = (meta.get('term_loc') or '').strip()
    if orig and term and orig != term:
        return f"{orig}->{term}"
    cable = (meta.get('cable_id') or '').strip()
    if cable:
        return cable
    return '?'


def load_fibers(input_path: str, direction: str = None) -> tuple:
    """Load every fiber from input_path.  Returns (fibers, chosen_direction).

    Direction is determined from each file's GenParams metadata (SOR) or
    GeneralParameters block (JSON), NOT from the filename.

    If ``direction`` is supplied, only files whose signature matches are
    kept.  Otherwise:
      * one signature in the data → use it all
      * many signatures → auto-pick the most populous, warn the user.
    """
    raw = []   # list of (fnum, filepath, name, parsed_dict, signature)
    for filepath, name in _walk_files(input_path):
        try:
            if name.lower().endswith('.json'):
                r = parse_otdr_json(filepath)
                meta = _read_json_genparams(filepath)
                fmt = 'json'
            else:
                r = parse_sor_full(filepath, trim=False)
                meta = _read_sor_genparams(filepath)
                fmt = 'sor'
        except Exception as exc:
            print(f"  WARN: parse failed for {name}: {exc}")
            continue
        if not r:
            continue
        # Prefer GenParams fiber_id ("0001" / "0144" — clean) over filename
        # heuristics, which can be fooled by numeric prefixes (CLE1CLE2…).
        fnum = None
        fiber_id = (meta or {}).get('fiber_id') or ''
        if fiber_id.strip():
            try:
                fnum = int(fiber_id.strip())
            except ValueError:
                fnum = None
        if fnum is None:
            fnum = _extract_fiber_num(name)
        if fnum is None:
            continue
        r['_source']      = fmt
        r['_fiber_num']   = fnum
        r['_genparams']   = meta
        r['_direction']   = direction_signature(meta)
        raw.append((fnum, filepath, name, r, r['_direction']))

    if not raw:
        return {}, None

    sig_counts = Counter(s for _, _, _, _, s in raw)
    print("  Direction signatures detected (from file metadata):")
    for sig, cnt in sig_counts.most_common():
        print(f"    {sig!r:40s}  {cnt} fibers")

    if direction is not None:
        chosen = direction
        if chosen not in sig_counts:
            print(f"  ERROR: requested --direction {chosen!r} not present.")
            return {}, None
    elif len(sig_counts) == 1:
        chosen = next(iter(sig_counts))
    else:
        chosen = sig_counts.most_common(1)[0][0]
        print(f"  Multiple directions present; auto-selected the most populous: {chosen!r}")
        print("  (override with --direction <signature>)")

    fibers = {}
    for fnum, filepath, name, r, sig in raw:
        if sig != chosen:
            continue
        if fnum in fibers:
            print(f"  WARN: duplicate fiber {fnum} in direction {sig!r} ({name}); keeping first")
            continue
        fibers[fnum] = r
    return fibers, chosen


# ═══════════════════════════════════════════════════════════════════════
#  STEP 2 — Normalize untrimmed traces (strip OTDR port + far-end conn)
# ═══════════════════════════════════════════════════════════════════════

def _normalize_untrimmed_events(events):
    """Match splicereportmatchexfo's _normalize_untrimmed_events: detect the
    OTDR port + launch pattern and re-reference distances to the launch
    connector.  Returns a new list (does not mutate input)."""
    if len(events) < 3:
        return events

    e0, e1 = events[0], events[1]
    if not (e0['is_reflective'] and not e0['is_end'] and
            e0['time_of_travel'] == 0 and
            e1['is_reflective'] and not e1['is_end'] and
            0 < e1['dist_km'] < LAUNCH_FIBER_MAX):
        return events

    launch_dist = e1['dist_km']
    launch_travel = e1['time_of_travel']

    end_idx = None
    for i, e in enumerate(events):
        if e['is_end']:
            end_idx = i
            break

    far_end_idx = None
    if end_idx is not None and end_idx > 1:
        end_dist = events[end_idx]['dist_km']
        for i in range(end_idx - 1, 0, -1):
            if events[i]['is_reflective'] and not events[i]['is_end']:
                if (end_dist - events[i]['dist_km']) < LAUNCH_FIBER_MAX:
                    far_end_idx = i
                break

    far_end_norm_dist = None
    if far_end_idx is not None:
        far_end_norm_dist = round(events[far_end_idx]['dist_km'] - launch_dist, 4)

    normalized = []
    for i, e in enumerate(events):
        if i == 0:
            continue
        if i == far_end_idx:
            continue
        new_e = dict(e)
        new_e['dist_km'] = round(e['dist_km'] - launch_dist, 4)
        new_e['time_of_travel'] = max(0, e['time_of_travel'] - launch_travel)
        if e['is_end'] and far_end_norm_dist is not None:
            new_e['dist_km'] = far_end_norm_dist
        normalized.append(new_e)
    return normalized


def normalize_all(fibers: dict):
    for fnum, r in fibers.items():
        r['events'] = _normalize_untrimmed_events(r['events'])


# ═══════════════════════════════════════════════════════════════════════
#  STEP 3 — Discover splice positions from the A-direction population
# ═══════════════════════════════════════════════════════════════════════

def discover_splices(fibers: dict):
    bins = defaultdict(list)
    for r in fibers.values():
        for e in r['events']:
            if e['dist_km'] < 1.0 or e['is_end']:
                continue
            t = e.get('type') or ''
            if not (t.startswith('0F') or t.startswith('1F')):
                continue
            bk = round(e['dist_km'])
            bins[bk].append(e['dist_km'])

    splices = []
    for bk in sorted(bins.keys()):
        if len(bins[bk]) < MIN_POP_SPLICE:
            continue
        avg_pos = round(np.mean(bins[bk]), 2)
        splices.append({'bin': bk, 'position_km': avg_pos, 'count': len(bins[bk])})

    merged = []
    for sp in splices:
        if merged and abs(sp['position_km'] - merged[-1]['position_km']) < 1.0:
            if sp['count'] > merged[-1]['count']:
                merged[-1] = sp
        else:
            merged.append(sp)

    return merged


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4 — Refine + validate closures (drop bend/damage phantoms)
# ═══════════════════════════════════════════════════════════════════════

def refine_and_validate(fibers: dict, splices):
    out = []
    dropped = []
    for sp in splices:
        center_guess = sp['position_km']
        nearby_pos, nearby_loss = [], []
        for r in fibers.values():
            for e in r['events']:
                if e['dist_km'] < 1.0 or e['is_end']:
                    continue
                if abs(e['dist_km'] - center_guess) < 1.0:
                    nearby_pos.append(e['dist_km'])
                    nearby_loss.append(e.get('splice_loss') or 0.0)

        if not nearby_pos:
            sp['position_km_refined'] = center_guess
            dropped.append(sp)
            continue

        arr = np.array(nearby_pos)
        bin_km = CLOSURE_MODE_BIN_M / 1000.0
        nbins = max(5, int(np.ceil((arr.max() - arr.min()) / bin_km)))
        hist, edges = np.histogram(arr, bins=nbins, range=(arr.min(), arr.max()))
        peak_idx = int(np.argmax(hist))
        peak_center = (edges[peak_idx] + edges[peak_idx + 1]) / 2.0
        local_mask = np.abs(arr - peak_center) < (CLOSURE_MODE_WINDOW_M / 1000.0)
        if local_mask.sum() >= 5:
            refined = float(np.median(arr[local_mask]))
        else:
            refined = float(peak_center)
        sp['position_km_refined'] = refined

        # Loss-distribution phantom rejection
        loss_arr = np.array(nearby_loss)
        tight_mask = np.abs(arr - refined) < CLOSURE_MATCH_KM
        tight_losses = loss_arr[tight_mask]
        is_phantom = False
        if len(tight_losses) >= MIN_POP_SPLICE:
            gainer_frac = float((tight_losses < 0).sum()) / len(tight_losses)
            median_loss = float(np.median(tight_losses))
            if gainer_frac < CLOSURE_VALID_MIN_GAINER_FRAC and median_loss > CLOSURE_VALID_MEDIAN_LOSS_MAX:
                is_phantom = True
                sp['phantom_reason'] = (
                    f'gainers={gainer_frac:.2f}, median_loss={median_loss:+.3f}dB'
                )

        # Display-km: lowest-numbered fiber with an event within ±CLOSURE_MATCH_KM
        # of the refined center, distance truncated to 10 m.
        display_km = refined
        for fnum in sorted(fibers.keys()):
            r = fibers[fnum]
            for e in r['events']:
                if e.get('is_end') or e['dist_km'] < 1.0:
                    continue
                if abs(e['dist_km'] - refined) < CLOSURE_MATCH_KM:
                    display_km = e['dist_km']
                    break
            if display_km != refined:
                break
        sp['position_km_display'] = math.floor(display_km * 100) / 100.0

        if is_phantom:
            dropped.append(sp)
        else:
            out.append(sp)

    if dropped:
        print(f"  Dropped {len(dropped)} phantom closure(s):")
        for sp in dropped:
            reason = sp.get('phantom_reason', 'no_data')
            print(f"    {sp['position_km']:8.2f} km  → {reason}")

    return out


# ═══════════════════════════════════════════════════════════════════════
#  STEP 5 — Find off-splice events
# ═══════════════════════════════════════════════════════════════════════

def find_off_splice_events(fibers: dict, valid_splices):
    """Return list of dicts: {fiber, position_km, loss}.

    An event is off-splice if:
      • it's a valid 0F/1F event (not end-of-fiber)
      • it's past the launch zone (>= LAUNCH_FIBER_MAX km from fiber start)
      • it's at least END_REGION_KM km before the fiber end
      • |loss| >= BEND_THRESHOLD
      • it's NOT within ±CLOSURE_MATCH_KM of any valid splice center
    """
    centers = [sp['position_km_refined'] for sp in valid_splices]

    off_events = []
    for fnum, r in fibers.items():
        events = r['events']
        end_km = None
        for e in events:
            if e.get('is_end'):
                end_km = e['dist_km']
                break

        for e in events:
            if e.get('is_end'):
                continue
            t = e.get('type') or ''
            if not (t.startswith('0F') or t.startswith('1F')):
                continue
            pos = e['dist_km']
            if pos < LAUNCH_FIBER_MAX:
                continue
            if end_km is not None and pos > (end_km - END_REGION_KM):
                continue
            loss = e.get('splice_loss') or 0.0
            if abs(loss) < BEND_THRESHOLD:
                continue
            # Skip if near any valid splice
            if centers and any(abs(pos - c) < CLOSURE_MATCH_KM for c in centers):
                continue
            off_events.append({
                'fiber': fnum,
                'position_km': pos,
                'loss': loss,
            })

    return off_events


def cluster_off_splice(off_events, fibers: dict):
    """Group off-splice events (≥ BEND_THRESHOLD = 0.1 dB) into columns.

    Every event within ``OFF_SPLICE_CLUSTER_M`` (100 m) of another is
    merged into the same cluster.  Whether a cluster has 1 fiber or 50,
    it receives the same tag: ``bend_damage`` — rendered yellow on the
    sheet.  Single-fiber and multi-fiber clusters are visually
    indistinguishable, which is what the tech wanted: "possible bend or
    damage" covers both.
    """
    if not off_events:
        return []

    sorted_evs = sorted(off_events, key=lambda x: x['position_km'])
    clusters = [[sorted_evs[0]]]
    tol_km = OFF_SPLICE_CLUSTER_M / 1000.0
    for ev in sorted_evs[1:]:
        if ev['position_km'] - clusters[-1][-1]['position_km'] <= tol_km:
            clusters[-1].append(ev)
        else:
            clusters.append([ev])

    columns = []
    for cl in clusters:
        positions = np.array([e['position_km'] for e in cl])
        refined = float(np.median(positions))
        lowest = min(cl, key=lambda e: e['fiber'])
        display_km = math.floor(lowest['position_km'] * 100) / 100.0
        columns.append({
            'position_km_refined': refined,
            'position_km_display': display_km,
            'members': cl,
            'kind': 'bend_damage',
        })

    columns.sort(key=lambda c: c['position_km_refined'])
    return columns


# ═══════════════════════════════════════════════════════════════════════
#  STEP 5.5 — Detect breaks (fiber dies before the cable end)
# ═══════════════════════════════════════════════════════════════════════

def find_breaks(fibers: dict, valid_splices, span_km: float):
    """Return list of break events: {fiber, position_km}.

    A fiber is broken if its end-of-fiber marker lies at least
    ``BREAK_PREMATURE_KM`` short of the cable span AND not within
    ``CLOSURE_MATCH_KM`` of any validated splice center.  Termination at
    a known closure is excluded — that's a fiber legitimately ending at
    a tap or pre-built segment, not a mid-span break.
    """
    if span_km <= 0:
        return []
    threshold = span_km - BREAK_PREMATURE_KM
    centers = [sp['position_km_refined'] for sp in valid_splices]
    breaks = []
    for fnum, r in fibers.items():
        eof_km = None
        for e in r['events']:
            if e.get('is_end'):
                eof_km = e['dist_km']
                break
        if eof_km is None or eof_km <= 1.0:
            continue
        if eof_km >= threshold:
            continue
        if any(abs(eof_km - c) < CLOSURE_MATCH_KM for c in centers):
            continue
        breaks.append({'fiber': fnum, 'position_km': eof_km})
    return breaks


def cluster_breaks(breaks):
    """Group break events within ``OFF_SPLICE_CLUSTER_M`` (100 m) into
    columns of kind 'break' (rendered red)."""
    if not breaks:
        return []
    sorted_brks = sorted(breaks, key=lambda b: b['position_km'])
    clusters = [[sorted_brks[0]]]
    tol_km = OFF_SPLICE_CLUSTER_M / 1000.0
    for b in sorted_brks[1:]:
        if b['position_km'] - clusters[-1][-1]['position_km'] <= tol_km:
            clusters[-1].append(b)
        else:
            clusters.append([b])
    columns = []
    for cl in clusters:
        positions = np.array([b['position_km'] for b in cl])
        refined = float(np.median(positions))
        lowest = min(cl, key=lambda b: b['fiber'])
        display_km = math.floor(lowest['position_km'] * 100) / 100.0
        columns.append({
            'kind': 'break',
            'position_km_refined': refined,
            'position_km_display': display_km,
            'members': cl,
        })
    columns.sort(key=lambda c: c['position_km_refined'])
    return columns


# ═══════════════════════════════════════════════════════════════════════
#  STEP 6 — Build column list (splices + off-splice clusters, by distance)
# ═══════════════════════════════════════════════════════════════════════

def build_columns(valid_splices, off_columns, break_columns=None):
    cols = []
    for sp in valid_splices:
        cols.append({
            'kind': 'splice',
            'position_km_refined': sp['position_km_refined'],
            'position_km_display': sp.get('position_km_display',
                                          sp['position_km_refined']),
            'fiber_count': sp.get('count', 0),
        })
    cols.extend(off_columns)
    if break_columns:
        cols.extend(break_columns)
    cols.sort(key=lambda c: c['position_km_refined'])
    return cols


# ═══════════════════════════════════════════════════════════════════════
#  STEP 7 — Map (ribbon, column) → has_event
# ═══════════════════════════════════════════════════════════════════════

def build_ribbon_grid(fibers: dict, columns, ribbon_size: int):
    """For each (ribbon_idx, col_idx) pair, return the list of
    ``(fiber_num, signed_loss_db)`` tuples — every fiber in that ribbon
    whose A-event lives inside that column's window.  For ``break``
    columns the loss value is ``None`` (the cell label drops the loss
    portion entirely; the break is the event itself)."""
    grid = defaultdict(list)
    for ci, col in enumerate(columns):
        center = col['position_km_refined']
        if col['kind'] == 'splice':
            window_km = CLOSURE_MATCH_KM
        else:
            window_km = OFF_SPLICE_CLUSTER_M / 1000.0
        for fnum, r in fibers.items():
            ribbon_idx = (fnum - 1) // ribbon_size

            # Break column — match the fiber's EOF event, not its loss events
            if col['kind'] == 'break':
                for e in r['events']:
                    if e.get('is_end') and abs(e['dist_km'] - center) <= window_km:
                        grid[(ribbon_idx, ci)].append((fnum, None))
                        break
                continue

            # Splice / bend_damage columns — match loss events
            best = None
            for e in r['events']:
                if e.get('is_end') or e['dist_km'] < 1.0:
                    continue
                t = e.get('type') or ''
                if not (t.startswith('0F') or t.startswith('1F')):
                    continue
                loss = e.get('splice_loss') or 0.0
                ev_abs = abs(loss)
                # Universal 0.1 dB threshold (uni one-shot rule)
                if ev_abs < BEND_THRESHOLD:
                    continue
                if abs(e['dist_km'] - center) <= window_km:
                    if best is None or ev_abs > best[0]:
                        best = (ev_abs, loss)
            if best is not None:
                grid[(ribbon_idx, ci)].append((fnum, best[1]))
    return grid


def _format_cell_label(entries):
    """Turn a list of (fnum, signed_loss_db) into a compact cell label.

    Lists every fiber by number (ascending) followed by the worst signed
    loss in the group.  Leading zero is dropped from the loss value to
    match the bidirectional report's convention.  When every entry's
    loss is ``None`` (i.e. break column), the loss portion is omitted —
    the column distance already says where the break is.

    Examples:
      F23 .180
      F23,F47 .220
      F1,F4,F7,F8,F9,F11,F12 .340
      F12,F19 broke         (loss=None for all entries)
    """
    if not entries:
        return ''
    fibers = ','.join(f"F{f}" for f, _ in sorted(entries, key=lambda fl: fl[0]))
    losses = [loss for _, loss in entries if loss is not None]
    if not losses:
        return f"{fibers} broke"
    worst_loss = max(losses)

    def _loss_str(v):
        sign = '-' if v < 0 else ''
        s = f"{abs(v):.3f}"
        if s.startswith('0.'):
            s = s[1:]
        return sign + s

    return f"{fibers} {_loss_str(worst_loss)}"


# ═══════════════════════════════════════════════════════════════════════
#  STEP 8 — Cable span auto-detect
# ═══════════════════════════════════════════════════════════════════════

def auto_detect_span(fibers: dict) -> float:
    eofs = []
    for r in fibers.values():
        for e in r['events']:
            if e.get('is_end'):
                eofs.append(e['dist_km'])
                break
    if not eofs:
        return 0.0
    arr = np.array(eofs)
    top_q = np.percentile(arr, 75)
    return float(np.median(arr[arr >= top_q]))


# ═══════════════════════════════════════════════════════════════════════
#  STEP 9 — Excel output
# ═══════════════════════════════════════════════════════════════════════

def ribbon_label(ri: int, ribbon_size: int, n_fibers: int) -> str:
    first = ri * ribbon_size + 1
    last = min(first + ribbon_size - 1, n_fibers)
    ribbon_num = ri + 1
    tube = ''
    if ri < 48:
        tube_letter = chr(ord('A') + ri // 2)
        tube_num = (ri % 2) + 1
        tube = f" ({tube_letter}{tube_num})"
    return f"Fiber {first}-{last} ({ribbon_num}){tube}"


def _flagged_event_rows(grid, columns, ribbon_size, n_fibers):
    """Flatten the ribbon grid into a per-event list.

    Yields dicts {fiber, ribbon, column_kind, column_label, column_km,
    loss, reason}.  Used to populate the 'Flagged Events' sheet so a tech
    can answer 'why is this cell shaded?' for any cell.
    """
    splice_n = bend_n = break_n = 0
    col_labels = []
    for col in columns:
        if col['kind'] == 'splice':
            splice_n += 1
            col_labels.append(f"Splice {splice_n}")
        elif col['kind'] == 'break':
            break_n += 1
            col_labels.append(f"Break {break_n}")
        else:
            bend_n += 1
            col_labels.append(f"Bend/Damage {bend_n}")

    rows = []
    for (ri, ci), entries in grid.items():
        col   = columns[ci]
        label = col_labels[ci]
        for fnum, loss in entries:
            ribbon = ri + 1
            if col['kind'] == 'splice':
                reason = (
                    f"At splice closure ({col['fiber_count']}-fiber population). "
                    f"|loss| {abs(loss):.3f} dB ≥ {BEND_THRESHOLD:.3f} dB "
                    "uni-flag threshold — re-burn candidate."
                )
            elif col['kind'] == 'break':
                reason = (
                    f"BREAK — fiber's trace terminates at {col['position_km_display']:.2f} km, "
                    f"more than {BREAK_PREMATURE_KM:.1f} km short of the cable end, "
                    "not at any validated splice closure.  Cable damage or fiber cut."
                )
            else:
                reason = (
                    f"Possible bend/damage: A-side event >= {BEND_THRESHOLD:.3f} dB "
                    "away from any validated splice closure.  Loss "
                    f"{abs(loss):.3f} dB."
                )
            rows.append({
                'fiber':         fnum,
                'ribbon':        ribbon,
                'column_label':  label,
                'column_kind':   col['kind'],
                'column_km':     col['position_km_display'],
                'loss':          loss,
                'reason':        reason,
            })
    rows.sort(key=lambda r: (r['fiber'], r['column_km']))
    return rows


def write_xlsx(grid, columns, n_fibers, ribbon_size, span_km, output_path,
               site_a='', site_b=''):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unidir Events"

    n_ribbons = (n_fibers + ribbon_size - 1) // ribbon_size
    n_cols = len(columns)

    # ── Styles (Calibri 12 across the board, per tech direction) ──
    FONT_NAME = 'Calibri'
    FONT_SIZE = 12
    hdr_font       = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color="FFFFFF")
    hdr_fill_sp    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_fill_bend  = PatternFill(start_color="B7950B", end_color="B7950B", fill_type="solid")
    hdr_fill_break = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    a_km_font      = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color="1F4E79")
    b_km_font      = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color="8B0000")
    ribbon_font    = Font(name=FONT_NAME, size=FONT_SIZE)

    splice_shade   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    bend_shade     = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")  # yellow
    break_shade    = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")  # red
    break_text     = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color="FFFFFF")

    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # ── Header rows: B→A on row 1, A→B on row 2 (closer to the data,
    #    since the A end is where the OTDR shot from) ─────────────────
    ab_label = f"{site_a}→{site_b}:" if (site_a and site_b) else "A→B:"
    ba_label = f"{site_b}→{site_a}:" if (site_a and site_b) else "B→A:"
    ws.cell(row=1, column=1, value=ba_label).font = b_km_font
    ws.cell(row=2, column=1, value=ab_label).font = a_km_font
    for ci, col in enumerate(columns):
        excel_col = ci + 2
        km_a = col['position_km_display']
        km_b = math.floor((span_km - km_a) * 100) / 100.0 if span_km > 0 else 0.0
        ft_a = km_a * 3280.84
        ft_b = km_b * 3280.84
        cb = ws.cell(row=1, column=excel_col, value=f"{km_b:.2f}km / {ft_b:,.0f}ft")
        cb.font = b_km_font
        cb.alignment = Alignment(horizontal='center')
        ca = ws.cell(row=2, column=excel_col, value=f"{km_a:.2f}km / {ft_a:,.0f}ft")
        ca.font = a_km_font
        ca.alignment = Alignment(horizontal='center')

    # ── Row 3: column-type header ────────────────────────────────────
    ws.cell(row=3, column=1, value="Ribbon").font = hdr_font
    ws.cell(row=3, column=1).fill = hdr_fill_sp
    splice_n = bend_n = break_n = 0
    for ci, col in enumerate(columns):
        excel_col = ci + 2
        if col['kind'] == 'splice':
            splice_n += 1
            label = f"Splice {splice_n}"
            fill = hdr_fill_sp
        elif col['kind'] == 'break':
            break_n += 1
            label = f"Break {break_n}"
            fill = hdr_fill_break
        else:
            bend_n += 1
            label = f"Bend/Damage {bend_n}"
            fill = hdr_fill_bend
        c = ws.cell(row=3, column=excel_col, value=label)
        c.font = hdr_font
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

    # ── Data rows ────────────────────────────────────────────────────
    cell_text_font = Font(name=FONT_NAME, size=FONT_SIZE)
    for ri in range(n_ribbons):
        excel_row = ri + 4
        ws.cell(row=excel_row, column=1,
                value=ribbon_label(ri, ribbon_size, n_fibers)).font = ribbon_font
        for ci, col in enumerate(columns):
            excel_col = ci + 2
            cell = ws.cell(row=excel_row, column=excel_col)
            cell.border = border
            entries = grid.get((ri, ci))
            if entries:
                if col['kind'] == 'splice':
                    cell.fill = splice_shade
                    cell.font = cell_text_font
                elif col['kind'] == 'break':
                    cell.fill = break_shade
                    cell.font = break_text
                else:
                    cell.fill = bend_shade
                    cell.font = cell_text_font
                cell.value = _format_cell_label(entries)
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)

    # ── Legend sheet ────────────────────────────────────────────────
    leg = wb.create_sheet("Legend")
    leg.column_dimensions['A'].width = 18
    leg.column_dimensions['B'].width = 80
    rows = [
        ("Blue header",      "1F4E79", "FFFFFF",
         "Splice column — closure position discovered from A-side population "
         f"(>= {MIN_POP_SPLICE} fibers in a 1 km bin, mode-refined, validated)."),
        ("Lt. Blue cell",    "BDD7EE", "1F4E79",
         f"Ribbon has at least one fiber with |loss| >= {BEND_THRESHOLD:.3f} dB "
         f"within ±{int(CLOSURE_MATCH_KM*1000)} m of the splice center."),
        ("Gold header",      "B7950B", "FFFFFF",
         f"Possible Bend / Damage column — A-side event(s) >= {BEND_THRESHOLD:.3f} dB "
         f"clustered within {OFF_SPLICE_CLUSTER_M} m of each other, NOT within "
         f"±{int(CLOSURE_MATCH_KM*1000)} m of any validated splice.  One or more "
         "fibers; visual treatment is identical for single- and multi-fiber."),
        ("Yellow cell",      "FFEB3B", "5D4037",
         f"Ribbon has at least one fiber with a possible bend/damage event here."),
        ("Dark Red header",  "C00000", "FFFFFF",
         f"Break column — fiber's trace dies more than {BREAK_PREMATURE_KM:.1f} km "
         "short of the cable end AND not at any validated splice.  Cable cut, "
         "crush, or fiber damage."),
        ("Red cell",         "FF4444", "FFFFFF",
         "Ribbon has at least one broken fiber that terminates at this distance."),
    ]
    leg.cell(row=1, column=1, value="Color").font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)
    leg.cell(row=1, column=2, value="Meaning").font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)
    for i, (name, fc, tc, desc) in enumerate(rows, start=2):
        c = leg.cell(row=i, column=1, value=name)
        c.fill = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
        c.font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color=tc)
        leg.cell(row=i, column=2, value=desc).font = Font(name=FONT_NAME, size=FONT_SIZE)

    # ── Cell label format key ────────────────────────────────────────
    base = 2 + len(rows) + 1
    leg.cell(row=base, column=1, value="Cell label format").font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE)
    leg.cell(row=base, column=2,
             value="Each shaded cell shows the fiber(s) and worst-case loss (dB) "
                   "for that ribbon × column.  Leading zero is dropped from loss "
                   "values (e.g. .180 = 0.180 dB, –.025 = –0.025 dB).").font = Font(name=FONT_NAME, size=FONT_SIZE)
    label_rows = [
        ("F23 .180",
         "Single fiber.  Fiber 23 has an event here with loss 0.180 dB."),
        ("F23,F47 .220",
         "Two fibers, comma-separated, ascending.  Loss shown is the worst "
         "(most positive) of the group."),
        ("F1,F4,F7,F8,F9 .340",
         "All fibers in the ribbon with a flagged event at this column are "
         "listed.  No count abbreviation — fiber numbers are always shown."),
        ("F23 -.105",
         "Negative loss = apparent gainer (MFD mismatch at a splice between "
         "dissimilar fibers).  Shown signed so gainers stand out."),
    ]
    for i, (lbl, desc) in enumerate(label_rows, start=base + 1):
        c = leg.cell(row=i, column=1, value=lbl)
        c.font = Font(name='Courier New', size=FONT_SIZE)
        leg.cell(row=i, column=2, value=desc).font = Font(name=FONT_NAME, size=FONT_SIZE)

    # ── Column widths ────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 30
    for ci in range(n_cols):
        col_letter = openpyxl.utils.get_column_letter(ci + 2)
        ws.column_dimensions[col_letter].width = 26
    # Row height bump so wrapped Calibri 12 labels stay readable
    for ri in range(4, n_ribbons + 4):
        ws.row_dimensions[ri].height = 32
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = 'B4'

    # ── Flagged Events sheet (per-fiber detail) ─────────────────────
    ev_sheet = wb.create_sheet("Flagged Events")
    ev_hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ev_hdr_font = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color="FFFFFF")
    ev_row_font = Font(name=FONT_NAME, size=FONT_SIZE)
    headers = ['Fiber', 'Ribbon', 'Column', 'Distance (km)',
               'Loss (dB)', 'Kind', 'Why flagged']
    widths  = [10, 11, 20, 16, 12, 22, 90]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ev_sheet.cell(row=1, column=i, value=h)
        c.fill = ev_hdr_fill
        c.font = ev_hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        ev_sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    kind_fill = {
        'splice':       splice_shade,
        'bend_damage':  bend_shade,
        'break':        break_shade,
    }
    kind_label = {
        'splice':       'Splice',
        'bend_damage':  'Possible Bend/Damage',
        'break':        'BREAK',
    }

    rows = _flagged_event_rows(grid, columns, ribbon_size, n_fibers)
    for i, r in enumerate(rows, start=2):
        ev_sheet.cell(row=i, column=1, value=r['fiber']).font = ev_row_font
        ev_sheet.cell(row=i, column=2, value=r['ribbon']).font = ev_row_font
        ev_sheet.cell(row=i, column=3, value=r['column_label']).font = ev_row_font
        ev_sheet.cell(row=i, column=4, value=round(r['column_km'], 3)).font = ev_row_font
        if r['loss'] is None:
            loss_cell = ev_sheet.cell(row=i, column=5, value='broke')
            loss_cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color='C00000')
        else:
            loss_cell = ev_sheet.cell(row=i, column=5, value=round(r['loss'], 3))
            loss_cell.font = ev_row_font
            if abs(r['loss']) >= HIGH_LOSS_STANDALONE_DB:
                loss_cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color='C00000')
        kind_cell = ev_sheet.cell(row=i, column=6, value=kind_label[r['column_kind']])
        kind_cell.font = ev_row_font
        if r['column_kind'] == 'break':
            kind_cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color='FFFFFF')
        kind_cell.fill = kind_fill[r['column_kind']]
        ev_sheet.cell(row=i, column=7, value=r['reason']).font = Font(name=FONT_NAME, size=FONT_SIZE)
        ev_sheet.cell(row=i, column=7).alignment = Alignment(wrap_text=True, vertical='top')
    ev_sheet.freeze_panes = 'A2'
    # Auto-filter so the tech can sort/filter by fiber, ribbon, kind, or distance
    ev_sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"

    wb.save(output_path)
    print(f"  Saved: {output_path}  ({len(rows)} flagged-event rows)")


# ═══════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="A-direction-only event finder (splices + off-splice events) → Excel ribbon grid.")
    ap.add_argument('input', help="Directory or .zip containing .sor / .json files")
    ap.add_argument('--output', default='unidirectional_events.xlsx')
    ap.add_argument('--ribbon-size', type=int, default=RIBBON_SIZE)
    ap.add_argument('--direction', default=None,
                    help="Direction signature to filter to (from metadata, "
                         "e.g. 'CLE1CLE2' or 'TUL->BAR').  If omitted, the "
                         "most populous direction is auto-selected.")
    args = ap.parse_args()

    print(f"Loading fibers from: {args.input}")
    fibers, chosen_dir = load_fibers(args.input, direction=args.direction)
    if not fibers:
        print("ERROR: no SOR/JSON files found / no fibers in selected direction")
        sys.exit(1)
    print(f"  Loaded {len(fibers)} fibers (direction: {chosen_dir!r})")

    normalize_all(fibers)

    print("Discovering splice closures (A-side population)...")
    candidates = discover_splices(fibers)
    print(f"  {len(candidates)} candidate position(s) before validation")

    valid = refine_and_validate(fibers, candidates)
    print(f"  {len(valid)} valid splice column(s)")

    print("Finding off-splice events...")
    off_evs = find_off_splice_events(fibers, valid)
    off_cols = cluster_off_splice(off_evs, fibers)
    print(f"  {len(off_evs)} off-splice events → {len(off_cols)} bend/damage column(s)")

    n_fibers = max(fibers.keys())
    span = auto_detect_span(fibers)
    print(f"  Cable span ≈ {span:.2f} km")

    print("Detecting broken fibers...")
    breaks = find_breaks(fibers, valid, span)
    break_cols = cluster_breaks(breaks)
    print(f"  {len(breaks)} broken fiber(s) → {len(break_cols)} break column(s)")

    columns = build_columns(valid, off_cols, break_cols)

    grid = build_ribbon_grid(fibers, columns, args.ribbon_size)

    # Derive 3-letter site codes from the first fiber's metadata
    sample = next(iter(fibers.values()))
    meta = sample.get('_genparams') or {}
    site_a = _short_code(meta.get('orig_loc'))
    site_b = _short_code(meta.get('term_loc'))
    # Fall back to splitting the direction string when orig/term were
    # blank in metadata (cable_id-only signatures like 'CLE1CLE2').
    if (not site_a or not site_b) and chosen_dir and '->' in chosen_dir:
        a, b = chosen_dir.split('->', 1)
        site_a = site_a or _short_code(a)
        site_b = site_b or _short_code(b)
    if site_a and site_b:
        print(f"  Sites: A end = {site_a}  →  B end = {site_b}")

    print(f"Writing Excel ({(n_fibers + args.ribbon_size - 1)//args.ribbon_size} ribbons × {len(columns)} columns)...")
    write_xlsx(grid, columns, n_fibers, args.ribbon_size, span, args.output,
               site_a=site_a, site_b=site_b)


if __name__ == '__main__':
    main()
