"""
acquisition_audit.py — Acquisition-Parameter consistency audit
==============================================================

Reads the per-fiber records produced by the existing SOR / JSON parsers
(sor_reader324802a, json_reader) and reports whether every trace in the
run was shot with the same instrument and the same settings.

The audit lives at the START of every Excel report (sheet index 0,
selected as the active sheet so the workbook opens on it) and at the
start of any PDF.  It is additive — no existing analysis or sheet is
modified.

Fields it compares, per trace:
    • Test timestamp   (bucketed by calendar day)
    • OTDR model       (exact match)
    • OTDR serial      (exact match)
    • Wavelength       (set of wavelengths per file)
    • Pulse width      (ns; per wavelength)
    • Averaging        (count for SOR/TRC, time for JSON)

Rule, applied per field:
    • Identical across all traces → green row, shows the spec.
    • Differs                     → amber row, shows the majority value
                                    and its count, then lists every
                                    file that differs with what it has.

Author: Unidirectional One Shot, May 2026.
"""
from __future__ import annotations

import datetime as dt
import os
from collections import Counter
from typing import Iterable

try:
    import openpyxl                         # type: ignore
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:                         # pragma: no cover
    openpyxl = None                          # caller must check


# ─────────────────────────────────────────────────────────────────────
#  Helpers — small enough to keep inline
# ─────────────────────────────────────────────────────────────────────

def _seconds_to_ns(seconds):
    """Convert pulse width in seconds → integer nanoseconds.  None-safe."""
    if seconds is None:
        return None
    try:
        return int(round(float(seconds) * 1e9))
    except (TypeError, ValueError):
        return None


def _epoch_to_day(epoch):
    """Bucket a Unix epoch (int or float) into a YYYY-MM-DD date string.
    Returns None when ``epoch`` is missing / zero / not a number."""
    if not epoch:
        return None
    try:
        epoch = float(epoch)
    except (TypeError, ValueError):
        return None
    if epoch <= 0:
        return None
    try:
        return dt.datetime.utcfromtimestamp(epoch).strftime("%Y-%m-%d")
    except (OSError, ValueError, OverflowError):
        return None


def _iso_to_day(s):
    """Parse a JSON-style ``TestDateTime`` ISO-8601 string into a day.
    Tolerates ``Z`` suffix and missing seconds."""
    if not s or not isinstance(s, str):
        return None
    s = s.strip().rstrip("Z")
    # dt.fromisoformat handles "2026-04-20T23:50:20" cleanly on 3.7+
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f",
                "%Y-%m-%dT%H:%M",    "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Last resort
    try:
        return dt.datetime.fromisoformat(s).strftime("%Y-%m-%d")
    except ValueError:
        return None


def _epoch_to_iso(epoch):
    """Pretty timestamp for the audit ('full earliest → latest span')."""
    try:
        epoch = float(epoch)
    except (TypeError, ValueError):
        return None
    if epoch <= 0:
        return None
    return dt.datetime.utcfromtimestamp(epoch).strftime("%Y-%m-%d %H:%M:%S UTC")


# ─────────────────────────────────────────────────────────────────────
#  Per-fiber acquisition record
# ─────────────────────────────────────────────────────────────────────

def extract_acquisition(fiber: dict) -> dict:
    """Pull the audit fields out of a fiber dict produced by
    parse_sor_full / parse_otdr_json.  Returns a flat dict with these
    keys::

        filename         basename of the source SOR / JSON
        source           'sor' | 'json'  (from engine's _source)
        test_day         'YYYY-MM-DD'  | None
        test_iso         pretty timestamp | None
        otdr_model       str | None
        otdr_serial      str | None
        wavelength_nm    int (rounded) | None
        pulse_ns         int | None
        averaging_count  int | None  (SOR/TRC count of acquisitions)
        averaging_secs   float | None  (JSON time-based averaging)

    Caller is responsible for not crashing on missing fields.  None
    values are how we tell the audit "this file didn't supply it".
    """
    src    = (fiber.get('_source') or '').lower()
    fname  = fiber.get('filename') or os.path.basename(fiber.get('filepath') or '')
    rec = {
        'filename':        fname,
        'source':          src,
        'test_day':        None,
        'test_iso':        None,
        'otdr_model':      None,
        'otdr_serial':     None,
        'wavelength_nm':   None,
        'pulse_ns':        None,
        'averaging_count': None,
        'averaging_secs':  None,
    }

    if src == 'json':
        rec['test_day']  = _iso_to_day(fiber.get('_json_test_datetime'))
        rec['test_iso']  = (fiber.get('_json_test_datetime') or '').strip() or None
        rec['otdr_model']  = fiber.get('_json_otdr_model') or None
        rec['otdr_serial'] = fiber.get('_json_otdr_serial') or None
        wl = fiber.get('_json_wavelength_nm')
        rec['wavelength_nm']   = int(round(wl)) if wl else None
        pulse = fiber.get('_json_pulse_ns')
        rec['pulse_ns']        = int(round(pulse)) if pulse else None
        rec['averaging_secs']  = fiber.get('_json_averaging_seconds')
    else:  # SOR (or anything else that produces SOR-shape dict)
        rec['test_day']  = _epoch_to_day(fiber.get('date_time'))
        rec['test_iso']  = _epoch_to_iso(fiber.get('date_time'))
        sup = fiber.get('sup_params') or {}
        rec['otdr_model']  = sup.get('module_id')    or None
        rec['otdr_serial'] = (sup.get('module_sn')   or
                              sup.get('mainframe_sn') or None)
        wl = fiber.get('wavelength')
        if wl is None:
            cal = fiber.get('exfo_calibration') or {}
            wl  = cal.get('ExactWavelength') or cal.get('NominalWavelength')
        rec['wavelength_nm'] = int(round(wl)) if wl else None
        cal = fiber.get('exfo_calibration') or {}
        pulse_s = cal.get('CalibratedPulseWidth') or cal.get('NominalPulseWidth')
        rec['pulse_ns']        = _seconds_to_ns(pulse_s)
        n_avg = cal.get('NumberOfAverages')
        rec['averaging_count'] = int(n_avg) if n_avg else None
    return rec


# ─────────────────────────────────────────────────────────────────────
#  Consistency algorithm — majority + outliers, one helper for all fields
# ─────────────────────────────────────────────────────────────────────

def _format_value(value, unit: str = '') -> str:
    if value is None or value == '':
        return '(missing)'
    if unit:
        return f"{value} {unit}".strip()
    return str(value)


def consistency(records: Iterable[tuple], label: str = '',
                unit: str = '') -> dict:
    """Given ``records`` = iterable of ``(filename, value)`` tuples, return::

        {
          'label':     label,
          'unit':      unit,
          'total':     N,
          'all_match': bool,
          'majority':  value,           # may be None when every value is None
          'majority_count': int,
          'all_missing':    bool,
          'outliers':  [(filename, value), ...],
          'spec':      formatted majority for display
        }

    Counting rules:
      • A ``None`` value counts as an outlier UNLESS every record is None
        (in which case the whole field is "Not available").
      • Ties on majority: the first-encountered most-common value wins.
        Order is deterministic because Counter.most_common is stable.
    """
    recs = list(records)
    total = len(recs)
    values = [v for _, v in recs]
    non_null = [v for v in values if v is not None and v != '']
    all_missing = len(non_null) == 0

    if all_missing or total == 0:
        return {
            'label': label, 'unit': unit, 'total': total,
            'all_match': False, 'majority': None, 'majority_count': 0,
            'all_missing': True, 'outliers': [],
            'spec': 'Not available (not stored in this file type)',
        }

    counts = Counter(non_null)
    majority, majority_count = counts.most_common(1)[0]
    outliers = [(fn, v) for fn, v in recs if v != majority]
    return {
        'label': label, 'unit': unit, 'total': total,
        'all_match': len(outliers) == 0,
        'majority': majority, 'majority_count': majority_count,
        'all_missing': False, 'outliers': outliers,
        'spec': _format_value(majority, unit),
    }


# ─────────────────────────────────────────────────────────────────────
#  Whole-audit builder
# ─────────────────────────────────────────────────────────────────────

def build_audit(fibers: dict) -> dict:
    """Build the complete audit dict from the engine's loaded fibers map.

    Structure::

      {
        'n_traces':      int,
        'file_level':    [ consistency_dict, ... ]
                         (model, serial, test_day, wavelengths)
        'date_span':     'YYYY-MM-DD HH:MM:SS UTC → YYYY-MM-DD HH:MM:SS UTC'
        'per_wavelength': {
            1550: { 'pulse': consistency_dict, 'avg': consistency_dict, 'n': int },
            1310: { ... },
          }
        'records':       [extract_acquisition(...) for each fiber]
      }
    """
    recs = [extract_acquisition(f) for f in fibers.values()]

    # ── File-level fields ───────────────────────────────────────────
    model_pairs  = [(r['filename'], r['otdr_model'])  for r in recs]
    serial_pairs = [(r['filename'], r['otdr_serial']) for r in recs]
    day_pairs    = [(r['filename'], r['test_day'])    for r in recs]
    wl_pairs     = [(r['filename'], r['wavelength_nm']) for r in recs]

    file_level = [
        consistency(model_pairs,  label='OTDR model'),
        consistency(serial_pairs, label='OTDR serial'),
        consistency(day_pairs,    label='Test date (calendar day, UTC)'),
        consistency(wl_pairs,     label='Wavelength', unit='nm'),
    ]

    # ── Date span (earliest → latest) ──────────────────────────────
    isos = [r['test_iso'] for r in recs if r['test_iso']]
    date_span = None
    if isos:
        # Both SOR (epoch-derived) and JSON (ISO) iso strings sort
        # lexicographically because they share YYYY-MM-DD prefix.
        isos_sorted = sorted(isos)
        date_span = f"{isos_sorted[0]} → {isos_sorted[-1]}"

    # ── Per-wavelength: pulse + averaging ──────────────────────────
    by_wl: dict = {}
    for r in recs:
        wl = r['wavelength_nm'] or 'unknown'
        by_wl.setdefault(wl, []).append(r)

    per_wavelength: dict = {}
    for wl, group in by_wl.items():
        pulse_pairs = [(r['filename'], r['pulse_ns']) for r in group]
        # Use whichever averaging field is non-null per record; record
        # the source for the spec label.
        avg_pairs_count = [(r['filename'], r['averaging_count']) for r in group]
        avg_pairs_secs  = [(r['filename'], r['averaging_secs'])  for r in group]
        non_null_count = sum(1 for _, v in avg_pairs_count if v is not None)
        non_null_secs  = sum(1 for _, v in avg_pairs_secs  if v is not None)
        if non_null_count >= non_null_secs:
            avg_consistency = consistency(avg_pairs_count, label='Averaging',
                                           unit='avg')
        else:
            avg_consistency = consistency(avg_pairs_secs, label='Averaging',
                                           unit='s')

        per_wavelength[wl] = {
            'n':     len(group),
            'pulse': consistency(pulse_pairs, label='Pulse width', unit='ns'),
            'avg':   avg_consistency,
        }

    return {
        'n_traces':     len(recs),
        'file_level':   file_level,
        'date_span':    date_span,
        'per_wavelength': per_wavelength,
        'records':      recs,
    }


# ─────────────────────────────────────────────────────────────────────
#  Excel writer — inserts the audit sheet at index 0
# ─────────────────────────────────────────────────────────────────────

# ── Colours pulled from the existing Excel report's house style ──
_FONT_NAME      = "Calibri"
_FONT_SIZE      = 12
_FG_WHITE       = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=True, color="FFFFFF") if openpyxl else None
_FG_DARK        = Font(name=_FONT_NAME, size=_FONT_SIZE) if openpyxl else None
_FG_BOLD        = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=True) if openpyxl else None
_FG_DIM         = Font(name=_FONT_NAME, size=_FONT_SIZE, italic=True,
                       color="6B6B6B") if openpyxl else None

_HDR_FILL       = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") if openpyxl else None
_GREEN_FILL     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if openpyxl else None
_AMBER_FILL     = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") if openpyxl else None
_AMBER_DIM_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") if openpyxl else None
_GREEN_FONT     = Font(name=_FONT_NAME, size=_FONT_SIZE, color="006100") if openpyxl else None
_AMBER_FONT     = Font(name=_FONT_NAME, size=_FONT_SIZE, color="9C5700") if openpyxl else None

_BORDER         = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
) if openpyxl else None


def _write_row(ws, row: int, label, result, value_override=None):
    """Render one consistency row at ``row``.  Returns the next row index
    (skipping any outlier child rows we emitted)."""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = _FG_BOLD
    label_cell.alignment = Alignment(vertical='top', wrap_text=True)
    label_cell.border = _BORDER

    value_cell = ws.cell(row=row, column=2)
    value_cell.alignment = Alignment(vertical='top', wrap_text=True)
    value_cell.border = _BORDER

    if result.get('all_missing'):
        value_cell.value = result['spec']
        value_cell.font  = _FG_DIM
        value_cell.fill  = _AMBER_DIM_FILL
        return row + 1

    if result['all_match']:
        text = value_override if value_override is not None else f"✓ All match: {result['spec']}"
        value_cell.value = text
        value_cell.font  = _GREEN_FONT
        value_cell.fill  = _GREEN_FILL
        return row + 1

    # Differing — majority + outliers
    header = (f"⚠ Majority: {result['spec']} ({result['majority_count']} of "
              f"{result['total']}) — {len(result['outliers'])} differ:")
    value_cell.value = header
    value_cell.font  = _AMBER_FONT
    value_cell.fill  = _AMBER_FILL

    next_row = row + 1
    for fn, v in result['outliers'][:64]:
        l = ws.cell(row=next_row, column=1, value=f"   {fn}")
        l.font = _FG_DIM
        l.alignment = Alignment(vertical='top', wrap_text=True)
        l.border = _BORDER
        rendered = _format_value(v, result['unit'])
        rcell = ws.cell(row=next_row, column=2, value=rendered)
        rcell.font = _FG_DIM
        rcell.fill = _AMBER_DIM_FILL
        rcell.alignment = Alignment(vertical='top', wrap_text=True)
        rcell.border = _BORDER
        next_row += 1
    if len(result['outliers']) > 64:
        more = ws.cell(row=next_row, column=2,
                       value=f"...and {len(result['outliers']) - 64} more")
        more.font = _FG_DIM
        more.fill = _AMBER_DIM_FILL
        next_row += 1
    return next_row


def write_audit_sheet(wb, audit: dict, sheet_title: str = "Acquisition Parameters") -> None:
    """Insert the audit as the FIRST sheet of ``wb`` and select it.

    Idempotent: if a sheet by ``sheet_title`` already exists (re-runs in
    the same session), it is removed and rewritten so the layout is
    always fresh.
    """
    if openpyxl is None:
        return
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]

    ws = wb.create_sheet(title=sheet_title, index=0)
    ws.sheet_properties.tabColor = "1F4E79"

    # ── Title row ───────────────────────────────────────────────────
    t = ws.cell(row=1, column=1, value="Acquisition consistency audit")
    t.font = Font(name=_FONT_NAME, size=14, bold=True, color="FFFFFF")
    t.fill = _HDR_FILL
    ws.cell(row=1, column=2).fill = _HDR_FILL
    sub = ws.cell(row=2, column=1,
                  value=f"{audit['n_traces']} traces in this run · "
                        f"green = all match, amber = mismatch with outliers listed below")
    sub.font = _FG_DIM
    ws.cell(row=2, column=2).font = _FG_DIM
    if audit.get('date_span'):
        span = ws.cell(row=3, column=1, value="Earliest → latest acquisition")
        span.font = _FG_BOLD
        span.border = _BORDER
        vsp = ws.cell(row=3, column=2, value=audit['date_span'])
        vsp.font  = _FG_DARK
        vsp.border = _BORDER

    # ── Section: file-level fields ─────────────────────────────────
    row = 5
    h = ws.cell(row=row, column=1, value="File-level fields (one value per trace)")
    h.font = _FG_WHITE
    h.fill = _HDR_FILL
    ws.cell(row=row, column=2).fill = _HDR_FILL
    row += 1
    for result in audit['file_level']:
        row = _write_row(ws, row, result['label'], result)

    # ── Section: per-wavelength fields ─────────────────────────────
    row += 1
    h = ws.cell(row=row, column=1,
                value="Per-wavelength fields (pulse width + averaging)")
    h.font = _FG_WHITE
    h.fill = _HDR_FILL
    ws.cell(row=row, column=2).fill = _HDR_FILL
    row += 1

    for wl in sorted(audit['per_wavelength'].keys(), key=lambda x: (x == 'unknown', x)):
        section = audit['per_wavelength'][wl]
        wl_label = (f"Wavelength {wl} nm" if wl != 'unknown'
                    else "Wavelength (missing in metadata)")
        l = ws.cell(row=row, column=1, value=f"{wl_label}  ({section['n']} traces)")
        l.font = _FG_BOLD
        l.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        l.border = _BORDER
        ws.cell(row=row, column=2).fill = PatternFill(
            start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        ws.cell(row=row, column=2).border = _BORDER
        row += 1
        row = _write_row(ws, row, "Pulse width",      section['pulse'])
        row = _write_row(ws, row, "Averaging",        section['avg'])
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 90
    ws.freeze_panes = "A4"

    # Make sure the workbook opens on the audit sheet
    wb.active = wb.sheetnames.index(sheet_title)
